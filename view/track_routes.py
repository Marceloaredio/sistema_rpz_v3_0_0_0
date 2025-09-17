# Rotas disponíveis para todos os usuários autenticados que possuem acesso ao módulo de Jornada.

from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
try:
    import pandas as pd
except ImportError:
    print("AVISO: pandas não disponível, usando stub")
    import pandas_stub as pd
import traceback
from datetime import datetime, timedelta, date
import sqlite3

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import json
import io
from io import BytesIO
import os

from controller.utils import convert_date_format, CustomLogger
from controller.google_sheets import GoogleSheetsManager
from controller.data import extract_data, allowed_file, fill_excel, fill_pdf, make_data_block
from controller.decorators import route_access_required
from controller.infractions import compute_infractions, convert_json_to_df
from controller.infractions_data import get_sorted_events_with_work_periods

from werkzeug.utils import secure_filename

from model.drivers.infractions_driver import InfractionsDriver
from model.drivers.track_analyzed_data_driver import AnalyzedTrackData
from model.drivers.motorist_driver import MotoristDriver
from model.drivers.uploaded_data_driver import UploadedDataDriver
from model.drivers.truck_driver import TruckDriver
from model.drivers.track_dayoff_driver import TrackDayOffDriver
from model.drivers.removed_infractions_driver import RemovedInfractionsDriver

from global_vars import DEBUG, DB_PATH, INFRACTION_DICT

track_bp = Blueprint('jornada', __name__)

routes_logger = CustomLogger(source="ROUTES", debug=DEBUG)
infraction_driver = InfractionsDriver(logger=routes_logger, db_path=DB_PATH)
motorist_driver = MotoristDriver(logger=routes_logger, db_path=DB_PATH)
uploaded_track_driver = UploadedDataDriver(logger=routes_logger, db_path=DB_PATH)
truck_driver = TruckDriver(logger=routes_logger, db_path=DB_PATH)
dayoff_driver = TrackDayOffDriver(logger=routes_logger, db_path=DB_PATH)
removed_infractions_driver = RemovedInfractionsDriver(logger=routes_logger, db_path=DB_PATH)
perm_uploaded_track_driver = AnalyzedTrackData(logger=routes_logger, db_path=DB_PATH)

def verify_conflicts(motorist_id: int, dates: list, db_conn) -> tuple:
    """
    Verifica se há conflitos de jornada para um motorista em uma lista de datas.

    Args:
        motorist_id: ID do motorista
        dates: Lista de datas no formato DD-MM-YYYY
        db_conn: Conexão com o banco de dados

    Returns:
        tuple: (nome do motorista, lista de conflitos)
    """
    # Inicializa a lista de conflitos
    conflicts = []
    cursor = db_conn.cursor()

    # Busca o nome do motorista com segurança
    try:
        cursor.execute("SELECT nome FROM motorists WHERE id = ?", (motorist_id,))
        row = cursor.fetchone()
        if not row or not isinstance(row, tuple) or len(row) < 1 or not row[0]:
            motorist_name = f"ID {motorist_id} (motorista não encontrado)"
        else:
            motorist_name = row[0]
    except Exception as e:
        motorist_name = f"ID {motorist_id} (erro ao buscar nome)"
        # AVISO: Erro ao buscar nome do motorista

    for date in dates:
        try:
            formatted_date = datetime.strptime(date, "%d-%m-%Y").strftime("%d/%m/%Y")
        except ValueError:
            formatted_date = date

        # Verifica se já tem jornada definida
        try:
            cursor.execute("SELECT truck_id FROM perm_data WHERE motorist_id = ? AND data = ?", (motorist_id, date))
            row = cursor.fetchone()
            if row:
                truck_id = row[0]
                cursor.execute("SELECT placa FROM trucks WHERE id = ?", (truck_id,))
                placa_row = cursor.fetchone()
                placa = placa_row[0] if placa_row and len(placa_row) > 0 else f"ID {truck_id}"
                conflicts.append(
                    f"Já existe jornada definida para o Motorista {motorist_name} com veículo {placa} na data '{formatted_date}'."
                )
        except Exception as e:
            # AVISO: Erro ao verificar perm_data
            pass

        # Verifica se tem folga marcada
        try:
            cursor.execute("SELECT motivo FROM dayoff WHERE motorist_id = ? AND data = ?", (motorist_id, date))
            row = cursor.fetchone()
            if row:
                motivo = row[0]
                conflicts.append(
                    f"Já existe jornada definida para o Motorista {motorist_name} com atribuição de '{motivo}' na data '{formatted_date}'."
                )
        except Exception as e:
            # AVISO: Erro ao verificar dayoff
            pass

    return motorist_name, conflicts

def generate_excel(tabela):
    """
    Gera um arquivo Excel com os dados da tabela.
    """
    import io
    from openpyxl import Workbook
    
    wb = Workbook()
    sheet = wb.active
    if not sheet:
        return None

    if not tabela:
        sheet.append(["Sem dados"])
    else:
        # Cabeçalhos
        cabecalhos = list(tabela[0].keys())
        sheet.append(cabecalhos)

        # Dados
        for linha in tabela:
            sheet.append([linha.get(col, "") for col in cabecalhos])

    # Salvar em memória
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

@track_bp.route('/mark_as_read', methods=['POST'])
@route_access_required
def mark_as_read():
    infraction_hash = request.form.get('hash')

    # Marca a infração como lida
    infraction_driver.mark_as_read(infraction_hash)

    return '', 204  # Retorna uma resposta vazia (204 No Content)

@track_bp.route('/upload', methods=['POST', 'GET'])
@route_access_required
def upload_file():
    if request.method == "POST":
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "Nenhum arquivo enviado"})

        file = request.files['file']
        tracker_type = request.form.get('tracker_type')
        truck_id = request.form.get('truck_id')

        if not file or not allowed_file(file.filename):
            return jsonify({"status": "error", "message": "Arquivo inválido"})

        try:
            filename = secure_filename(file.filename)

            # Definindo o diretório de destino para o tipo de rastreador
            if tracker_type == 'positron':
                upload_folder = os.path.join(os.getcwd(), 'raw_data', 'positron')
            elif tracker_type == 'sasgc':
                upload_folder = os.path.join(os.getcwd(), 'raw_data', 'sasgc')
            elif tracker_type == 'sascar':
                routes_logger.print("Salvando Sascar")
                upload_folder = os.path.join(os.getcwd(), 'raw_data', 'sascar')
            else:
                return jsonify({"status": "error", "message": "Tipo de rastreador inválido"})

            # Certifique-se de que o diretório existe
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)

            # Salvar o arquivo
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)

            routes_logger.print(f"Tracker Type: {tracker_type}")

            # Processamento do arquivo após o upload
            try:
                # Dependendo do tipo de rastreador, processa o arquivo
                if tracker_type == 'positron':
                    routes_logger.print("Extraindo Positron")
                    df = extract_data(filepath=file_path, system_type='positron')
                    df['truck_id'] = truck_id
                elif tracker_type == 'sasgc':
                    routes_logger.print("Extraindo Sasgc")
                    df = extract_data(filepath=file_path, system_type='sasgc')
                    df['truck_id'] = truck_id
                elif tracker_type == 'sascar':
                    routes_logger.print("Extraindo Sascar")
                    df = extract_data(filepath=file_path, system_type='sascar', truck_driver=truck_driver)

                # Insere os dados processados no banco de dados
                uploaded_track_driver.insert_from_dataframe(df)

                # Após o processamento, remove o arquivo
                try:
                    os.remove(file_path)
                except Exception as e:
                    routes_logger.register_log(f"Erro ao excluir arquivo {file_path}.", f"Erro: {e}")

                return jsonify({
                    "status": "success",
                    "message": f"Arquivo {filename} processado com sucesso!"
                })

            except Exception as e:
                routes_logger.register_log(f"Erro ao extrair dados do arquivo {file_path}.", f"Erro: {e}")
                return jsonify({
                    "status": "error",
                    "message": f"Erro ao processar arquivo {filename}: {str(e)}"
                })

        except Exception as e:
            routes_logger.register_log(f"Erro ao salvar arquivo.", f"Erro: {e}")
            return jsonify({
                "status": "error",
                "message": f"Erro ao salvar arquivo {file.filename}: {str(e)}"
            })

    # Para a renderização da página de upload, buscamos as placas e trucks
    all_trucks = truck_driver.retrieve_all_trucks()
    trucks_ids = [truck[0] for truck in all_trucks]
    plates = [truck[1] for truck in all_trucks]

    return render_template('load_files_track.html', plates=plates, trucks_ids=trucks_ids)

@track_bp.route('/clear_vehicle_data', methods=['POST'])
@route_access_required
def clear_vehicle_data():
    try:
        # Deleta todos os registros da tabela vehicle_data
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM vehicle_data")
        conn.commit()
        conn.close()
        
        return jsonify({"status": "success", "message": "Dados dos rastreadores limpos com sucesso!"})
    except Exception as e:
        routes_logger.register_log("Erro ao limpar dados dos rastreadores.", f"Erro: {e}")
        return jsonify({"status": "error", "message": "Erro ao limpar dados dos rastreadores."})

@track_bp.route('/track', methods=['GET'])
@route_access_required
def track():
    all_motorists = motorist_driver.retrieve_all_motorists()
    # Ordenar motoristas alfabeticamente por nome
    all_motorists.sort(key=lambda x: x[1].upper() if x[1] else '')
    all_motorists = [(motorist[0], motorist[1]) for motorist in all_motorists]

    available_trucks = uploaded_track_driver.get_unique_truck_ids_and_plates()

    return render_template('track.html', motorists=all_motorists, trucks=available_trucks)

@track_bp.route('/track_analysis')
@route_access_required
def track_analysis():
    truck_id = request.args.get('truck_id', '')
    motorist_id = request.args.get('motorist_id', '')

    if not truck_id or not motorist_id:
        flash("Placa/motorista não informada.")
        return render_template('redirect.html', url='/track')

    try:
        # Obtendo dados com a placa informada
        query_data = uploaded_track_driver.retrieve_all_records_by_condition(where_columns=['truck_id', ],
                                                                   where_values=(truck_id,))

        if len(query_data) == 0:
            flash("Sem dados para analisar sobre a placa informada. Insira novos dados através das configurações.")
            return render_template('redirect.html', url='/track')

        dates = set()

        for record in query_data:
            date_iso = record[1]  # índice 1 corresponde ao campo data_iso
            date_only = date_iso.split(" ")[0]
            dates.add(date_only)
        dates = sorted(dates)

        # Pega a data inicial e a data final
        initial_date = dates[0]
        formatted_initial_date = convert_date_format(initial_date)
        final_date = dates[-1]
        formatted_final_date = convert_date_format(final_date)

        blocks = []
        # Define as colunas do DataFrame
        colunas = [
            'truck_id',
            'data_iso',
            'vel',
            'latitude',
            'longitude',
            'uf',
            'cidade',
            'rua',
            'ignicao'
        ]

        for date in dates:
            # Registro de cada dia da placa especificada
            records = uploaded_track_driver.retrieve_by_datetime_range(start_datetime=date + " 00:00",
                                                             end_datetime=date + " 23:59",
                                                             where_columns=['truck_id'],
                                                             where_values=(truck_id,))

            # Cria o DataFrame vazio com essas colunas
            records_df = pd.DataFrame(records, columns=colunas)
            routes_logger.print(records_df)
            # DEBUG: Gerando blocos
            block = make_data_block(records_df, date)

            blocks.append(block)

        plate = truck_driver.retrieve_truck(where_columns=['id',],
                                            where_values=(truck_id,))[1]
        motorist_name = motorist_driver.retrieve_motorist(where_columns=['id',],
                                                          where_values=(motorist_id,))[1]
    except Exception as e:
        flash(f"Erro ao converter os dados em blocos para análise. Consulte o log ROUTER para maiores informações.")
        traceback.print_exc()
        routes_logger.register_log("Erro ao converter os dados em blocos para análise.", f"Erro: {e}")
        return render_template("redirect.html", url="/track")

    return render_template(
        "track_analysis.html",
        truck_id=truck_id,
        motorist_id=motorist_id,
        plate=plate,
        motorist=motorist_name,
        data_inicial=formatted_initial_date,
        data_final=formatted_final_date,
        blocos=blocks
    )


### Inserir dados
@track_bp.route('/insert_data', methods=['GET'])
@route_access_required
def insert_data():
    try:
        # Recupera todos os motoristas como tuplas (id, nome)
        all_motorists = motorist_driver.retrieve_all_motorists()
        # Ordenar motoristas alfabeticamente por nome
        all_motorists.sort(key=lambda x: x[1].upper() if x[1] else '')
        motorists_list = [(m[0], m[1]) for m in all_motorists if m and len(m) >= 2]



        return render_template(
            'track_insert_data_page.html',
            motorists=motorists_list
        )

    except Exception as e:
        routes_logger.register_log(f"[ERRO] insert_data: {e}")
        flash("Ocorreu um erro ao carregar a página de inserção.")
        return redirect(url_for('public.main_page'))
    

@track_bp.route('/save_dayoff', methods=['POST'])
@route_access_required
def save_dayoff():
    routes_logger.register_log(f"[DEBUG] save_dayoff: Iniciando função")
    routes_logger.register_log(f"[DEBUG] save_dayoff: request.form = {dict(request.form)}")
    
    if request.method == 'POST':
        motorist_id = request.form.get('motorist_id')
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        motivo = request.form.get('motivo')
        substituir = request.form.get('substituir', False)

        routes_logger.register_log(f"[DEBUG] save_dayoff: motorist_id={motorist_id}, start_date={start_date_str}, end_date={end_date_str}, motivo={motivo}")

        if not motorist_id:
            routes_logger.register_log(f"[DEBUG] save_dayoff: Erro - motorist_id não fornecido")
            return jsonify({"status": "error", "message": "Por favor, selecione um motorista."}), 400

        try:
            # Converter datas do formato YYYY-MM-DD para datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

            routes_logger.register_log(f"[DEBUG] save_dayoff: Datas convertidas - start_date={start_date}, end_date={end_date}")

            if end_date < start_date:
                routes_logger.register_log(f"[DEBUG] save_dayoff: Erro - Data de fim anterior à data de início")
                return jsonify({"status": "error", "message": "A Data de Fim deve ser igual ou posterior à Data de Início."}), 400

            # Gerar lista de datas no formato DD-MM-YYYY
            datas = []
            current_date = start_date
            while current_date <= end_date:
                datas.append(current_date.strftime('%d-%m-%Y'))
                current_date += timedelta(days=1)

            routes_logger.register_log(f"[DEBUG] save_dayoff: Datas geradas = {datas}")

            # Inicializar logger para INSERIR
            inserir_logger = CustomLogger(source="INSERIR", debug=DEBUG)
            
            # Recuperar nome do motorista
            motorista = motorist_driver.retrieve_motorist(['id'], (int(motorist_id),))
            routes_logger.register_log(f"[DEBUG] save_dayoff: Resultado busca motorista = {motorista}")
            
            if not motorista or not isinstance(motorista, tuple) or len(motorista) < 2:
                routes_logger.register_log(f"[DEBUG] save_dayoff: Erro - Dados do motorista inválidos")
                return jsonify({"status": "error", "message": "Erro ao recuperar dados do motorista."}), 400
            
            motorist_name = motorista[1]
            user_name = session.get('user', {}).get('name', 'Usuário desconhecido')

            routes_logger.register_log(f"[DEBUG] save_dayoff: motorist_name={motorist_name}, user_name={user_name}")

            # Preparar dados para verificação usando a nova lógica
            dados_dayoff = []
            for data in datas:
                dados_dayoff.append({
                    'Data': data,
                    'Observação': motivo
                })

            routes_logger.register_log(f"[DEBUG] save_dayoff: dados_dayoff preparados = {dados_dayoff}")

            # NOVA LÓGICA: Verificar conflitos ANTES de inserir qualquer coisa
            result = dayoff_driver.check_conflicts_only(dados_dayoff, int(motorist_id))

            # Debug: log dos conflitos encontrados
            routes_logger.register_log(f"[DEBUG] save_dayoff: motorist_id={motorist_id}, conflicts_found={result.get('tem_conflitos', False)}")
            routes_logger.register_log(f"[DEBUG] save_dayoff: result completo = {result}")

            # Se há conflitos e não é para substituir, retornar conflitos
            if result.get('tem_conflitos') and not substituir:
                routes_logger.register_log(f"[DEBUG] save_dayoff: Conflitos encontrados, retornando para frontend")
                # Preparar conflitos para compatibilidade com frontend
                conflitos_formatados = []
                for conflito in result.get('conflitos', []):
                    if isinstance(conflito, dict):
                        conflitos_formatados.append(f"{conflito['data']} - {conflito['tipo']}: {conflito['descricao']}")
                    else:
                        conflitos_formatados.append(str(conflito))
                
                # Retornar para o frontend decidir sobre substituição
                response_data = {
                    "status": "conflito",
                    "message": "Conflitos encontrados",
                    "conflitos": conflitos_formatados,
                    "conflitos_detalhados": result.get('conflitos', []),
                    "registros_inseridos": result.get('registros_inseridos', 0),
                    "registros_ignorados": result.get('registros_ignorados', 0)
                }
                return jsonify(response_data), 409

            # Se não há conflitos, inserir todos os registros de uma vez
            routes_logger.register_log(f"[DEBUG] save_dayoff: Iniciando inserção de dados")
            result = dayoff_driver.insert_data_from_json(dados_dayoff, int(motorist_id), replace=bool(substituir))
            
            routes_logger.register_log(f"[DEBUG] save_dayoff: Resultado da inserção = {result}")
            
            # Se chegou até aqui, inserção foi bem-sucedida
            routes_logger.register_log(f"[DEBUG] save_dayoff: Inserção concluída com sucesso")
            routes_logger.register_log(f"[DEBUG] save_dayoff: registros_inseridos={result.get('registros_inseridos', 0)}")
            routes_logger.register_log(f"[DEBUG] save_dayoff: registros_substituidos={result.get('registros_substituidos', 0)}")
            
            # Registrar no log
            inserir_logger.register_log(
                f"{'Substituído' if substituir else 'Adicionado'} com sucesso - Motorista: {motorist_name} (ID {motorist_id}) - "
                f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')} - "
                f"Motivo: {motivo} - Usuário responsável: {user_name}",
                None  # Não é um erro, apenas informação
            )
            
            routes_logger.register_log(f"[DEBUG] save_dayoff: Retornando sucesso")
            return jsonify({"status": "success", "message": "Dados salvos com sucesso!"})

        except ValueError as e:
            routes_logger.register_log(f"[DEBUG] save_dayoff: Erro ValueError = {e}")
            return jsonify({"status": "error", "message": "Formato de data inválido."}), 400

        except Exception as e:
            error_details = traceback.format_exc()
            routes_logger.register_log(f"[DEBUG] save_dayoff: Erro Exception = {e}")
            routes_logger.register_log(f"Erro ao inserir dados de folga: {e}", error_details)
            return jsonify({"status": "error", "message": "Ocorreu um erro ao inserir os dados."}), 500
        
@track_bp.route('/confirm_dayoff', methods=['POST'])
@route_access_required
def confirm_dayoff():
    routes_logger.register_log(f"[DEBUG] confirm_dayoff: Função iniciada")
    routes_logger.register_log(f"[DEBUG] confirm_dayoff: request.form = {dict(request.form)}")
    
    try:
        # Verifica se o botão "Sim" foi clicado no modal
        if 'confirmar' in request.form:
            routes_logger.register_log(f"[DEBUG] confirm_dayoff: Botão 'Sim' clicado")
            pendente = session.pop('pendente', None)

            if not pendente:
                routes_logger.register_log(f"[DEBUG] confirm_dayoff: Nenhum dado pendente encontrado")
                flash("Não havia dados pendentes para confirmar.")
                return redirect(url_for('jornada.insert_data'))

            motorist_id = int(pendente.get('motorist_id'))
            motivo = pendente.get('motivo')
            datas = pendente.get('datas')
            start_date_str = pendente.get('start_date')
            end_date_str = pendente.get('end_date')

            routes_logger.register_log(f"[DEBUG] confirm_dayoff: motorist_id={motorist_id}, motivo={motivo}, datas_count={len(datas) if datas else 0}")

            if not motorist_id or not datas or not motivo:
                routes_logger.register_log(f"[DEBUG] confirm_dayoff: Dados incompletos")
                flash("Dados incompletos para processar a substituição.")
                return redirect(url_for('jornada.insert_data'))

            # Recupera nome do motorista
            motorista = motorist_driver.retrieve_motorist(['id'], (motorist_id,))
            routes_logger.register_log(f"[DEBUG] Resultado da busca por motorista_id={motorist_id}: {motorista}")

            if not motorista or not isinstance(motorista, tuple) or len(motorista) < 2:
                flash("Erro ao recuperar nome do motorista.")
                return redirect(url_for('jornada.insert_data'))

            motorist_name = motorista[1]
            user_name = session.get('user', {}).get('name', 'Usuário desconhecido')

            inserir_logger = CustomLogger(source="INSERIR", debug=DEBUG)
            conn = sqlite3.connect(dayoff_driver.db_path)
            cursor = conn.cursor()

            # Contadores para o log final
            folgas_removidas = 0
            jornadas_removidas = 0

            for data in datas:
                # LOG antes da exclusão de dayoff
                cursor.execute("SELECT motivo FROM dayoff WHERE motorist_id = ? AND data = ?", (motorist_id, data))
                motivo_existente = cursor.fetchone()
                if motivo_existente:
                    inserir_logger.register_log(
                        f"Remoção de folga para {motorist_name} (ID {motorist_id}) na data {data} - motivo anterior '{motivo_existente[0]}' - Usuário responsável: {user_name}",
                        None  # Não é um erro, apenas informação
                    )
                    folgas_removidas += 1

                # LOG antes da exclusão de perm_data
                cursor.execute("SELECT truck_id FROM perm_data WHERE motorist_id = ? AND data = ?", (motorist_id, data))
                perm_existente = cursor.fetchone()
                if perm_existente:
                    inserir_logger.register_log(
                        f"Remoção de jornada para {motorist_name} (ID {motorist_id}) na data {data} - truck_id {perm_existente[0]} - Usuário responsável: {user_name}",
                        None  # Não é um erro, apenas informação
                    )
                    jornadas_removidas += 1

                # Remove registros antigos (perm_data e dayoff)
                cursor.execute("DELETE FROM perm_data WHERE motorist_id = ? AND data = ?", (motorist_id, data))
                cursor.execute("DELETE FROM dayoff WHERE motorist_id = ? AND data = ?", (motorist_id, data))

            conn.commit()
            conn.close()

            # Insere os novos dados
            for data in datas:
                dayoff_driver.replace_dayoff(motorist_id, data, motivo)

            routes_logger.register_log(f"[DEBUG] confirm_dayoff: Substituição concluída - {len(datas)} registros inseridos")

            # Log final de substituição bem-sucedida
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            
            inserir_logger.register_log(
                f"Substituído com sucesso - Motorista: {motorist_name} (ID {motorist_id}) - "
                f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')} - "
                f"Motivo: {motivo} - Folgas removidas: {folgas_removidas} - Jornadas removidas: {jornadas_removidas} - Usuário responsável: {user_name}",
                None  # Não é um erro, apenas informação
            )

            flash("Folgas substituídas com sucesso.")

        else:
            routes_logger.register_log(f"[DEBUG] confirm_dayoff: Botão 'Cancelar' clicado ou nenhum botão")
            # Se o usuário cancelou, registrar no log
            pendente = session.pop('pendente', None)
            if pendente:
                motorist_name = pendente.get('motorist_name', 'Motorista desconhecido')
                user_name = session.get('user', {}).get('name', 'Usuário desconhecido')
                
                inserir_logger = CustomLogger(source="INSERIR", debug=DEBUG)
                inserir_logger.register_log(
                    f"Conflito detectado - Ação cancelada pelo usuário - Motorista: {motorist_name} - Usuário responsável: {user_name}",
                    None  # Não é um erro, apenas informação
                )
            
            flash("Ação cancelada.")

    except Exception as e:
        error_details = traceback.format_exc()
        routes_logger.register_log("[ERRO] confirm_dayoff: erro interno ao processar substituição", error_details)
        flash("Erro interno ao confirmar substituição. Consulte o log para mais detalhes.")

    return redirect(url_for('jornada.insert_data'))

@track_bp.route('/export_dayoff_to_google', methods=['POST'])
@route_access_required
def export_dayoff_to_google():
    """
    Rota para exportar dados da tabela dayoff para o Google Sheets
    """
    try:
        from flask import jsonify
        
        dayoff_data = dayoff_driver.retrieve_all_dayoffs()
        
        if not dayoff_data:
            return jsonify({
                'success': False,
                'error': 'Nenhum dado de folga encontrado na tabela dayoff'
            })
        
        google_sheets_manager = GoogleSheetsManager(routes_logger)
        result = google_sheets_manager.export_dayoff_data(dayoff_data)
        return jsonify(result)
        
    except Exception as e:
        error_msg = f"Erro ao exportar dados dayoff para Google Sheets: {str(e)}"
        routes_logger.register_log(f"[ERRO] export_dayoff_to_google: {error_msg}")
        
        return jsonify({
            'success': False,
            'error': error_msg
        })

@track_bp.route('/check_updates')
@route_access_required
def check_updates_page():
    all_motorists = motorist_driver.retrieve_active_motorists_for_journey()
    # Ordenar motoristas alfabeticamente por nome
    all_motorists.sort(key=lambda x: x[1].upper() if x[1] else '')

    motorists_list = [{"id": m[0], "name": m[1]} for m in all_motorists]
    return render_template('check_updates.html', motorists=motorists_list)

### INFRAÇÕES

@track_bp.route('/infractions', methods=['GET'])
@route_access_required
def infractions_page():
    # Obtém os valores dos filtros da URL
    motorist_filter = request.args.get('motorist_id', '')
    tipo_infracao_filter = request.args.get('tipo_infracao', '')

    # Obter todos os motoristas disponíveis
    all_motorists = motorist_driver.retrieve_all_motorists()
    # Ordenar motoristas alfabeticamente por nome
    all_motorists.sort(key=lambda x: x[1].upper() if x[1] else '')
    all_motorists = [(motorist[0], motorist[1]) for motorist in all_motorists]

    # Construir as condições de filtro
    where_columns = []
    where_values = []
    
    if motorist_filter:
        where_columns.append('motorist_id')
        where_values.append(motorist_filter)
    
    if tipo_infracao_filter:
        where_columns.append('tipo_infracao')
        where_values.append(tipo_infracao_filter)

    # Se houver filtros, aplica-os, caso contrário retorna todas as infrações
    if where_columns and where_values:
        infractions = infraction_driver.retrieve_infractions(where_columns=where_columns,
                                                           where_values=tuple(where_values))
    else:
        infractions = infraction_driver.retrieve_all_infractions()

    # Passa as infrações, motoristas e tipos de infração para o template
    return render_template('infractions_page.html',
                         infractions=infractions,
                         all_motorists=all_motorists,
                         infraction_types=INFRACTION_DICT,
                         motorist_id=motorist_filter,
                         tipo_infracao=tipo_infracao_filter)

@track_bp.route('/infraction_details')
@route_access_required
def infraction_details():
    infraction_hash = request.args.get('hash')
    # DEBUG: Hash da infração
    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect('dbs/db_app.db')

    # Consulta SQL com chave composta (customer_id e order_id)
    query = f"""
        SELECT infractions.data, infractions.hora, infractions.duration, infractions.desc_infracao,
               infractions.link_tratativa,
               motorists.nome,
               trucks.placa,
               perm_data.*
        FROM infractions
        JOIN motorists ON infractions.motorist_id = motorists.id
        JOIN trucks ON infractions.truck_id = trucks.id
        JOIN perm_data 
        ON infractions.data = perm_data.data
        AND infractions.truck_id = perm_data.truck_id
        WHERE infractions.hash = '{infraction_hash}'
    """

    # Executar a consulta e armazenar o resultado em um DataFrame
    df = pd.read_sql_query(query, conn)

    # Fechar a conexão
    conn.close()

    data_dict = df.iloc[0].to_dict()

    return render_template('infractions_details.html', data_dict=data_dict)

@track_bp.route('/download_removed_infractions')
@route_access_required
def download_removed_infractions():
    try:
        # Buscar todas as infrações removidas
        removed_infractions = removed_infractions_driver.retrieve_all_removed_infractions()
        
        if not removed_infractions:
            flash("Nenhuma infração removida encontrada.")
            return redirect(url_for('jornada.infractions_page'))
        
        # Criar DataFrame
        df = pd.DataFrame(removed_infractions, columns=[
            'ID', 'Motorista', 'Placa', 'Data', 'Hora', 'Tipo Infração', 
            'Velocidade', 'Local', 'Data Remoção', 'Usuário Remoção'
        ])
        
        # Criar arquivo Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Infrações Removidas', index=False)
            
            # Formatar planilha
            workbook = writer.book
            worksheet = writer.sheets['Infrações Removidas']
            
            # Formatar cabeçalhos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        output.seek(0)
        
        # Nome do arquivo
        filename = f"infracoes_removidas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        routes_logger.register_log(f"[ERRO] download_removed_infractions: {e}")
        flash("Erro ao gerar relatório de infrações removidas.")
        return redirect(url_for('jornada.infractions_page'))
    

    
@track_bp.route('/remove_infraction', methods=['POST'])
@route_access_required
def remove_infraction():
    try:
        infraction_hash = request.args.get('hash')
        if not infraction_hash:
            return 'Hash não fornecido', 400

        # Busca os dados da infração antes de removê-la
        infraction_data = infraction_driver.retrieve_infraction(['hash'], (infraction_hash,))
        if not infraction_data:
            return 'Infração não encontrada', 404

        # Adiciona a infração à tabela de removidas
        removed_infractions_driver.add_removed_infraction(infraction_data)

        # Remove a infração da tabela principal
        infraction_driver.delete_infraction(['hash'], (infraction_hash,))

        return '', 204
    except Exception as e:
        routes_logger.register_log("Erro ao remover infração", str(e))
        return 'Erro ao remover infração', 500

@track_bp.route('/download-report', methods=['POST', 'GET'])
@route_access_required
def download_report():
    if request.method == "POST":
        file_name = None

        try:
            motorist_id = request.form.get('motorist_id')
            truck_id = request.form.get('truck_id')
            from_date = request.form.get('from_date')
            to_date = request.form.get('to_date')

            if not (motorist_id or truck_id) or not (from_date and to_date):
                flash("Selecione um Motorista e/ou Caminhão e uma data de início e fim.")
                return redirect("/download-report")

            if isinstance(from_date, str):
                from_date = from_date.replace('/', '-')
            else:
                flash("Ocorreu um erro durante a conversão de data.")
                return redirect("/download-report")

            if isinstance(to_date, str):
                to_date = to_date.replace('/', '-')
            else:
                flash("Ocorreu um erro durante a conversão de data.")
                return redirect("/download-report")

            query = "SELECT motorists.nome, trucks.placa, data, dia_da_semana, inicio_jornada, in_refeicao, fim_refeicao, fim_jornada, " \
                f"observacao, tempo_refeicao, intersticio, tempo_intervalo, tempo_carga_descarga, jornada_total, " \
                f"tempo_direcao, direcao_sem_pausa, in_descanso_1, fim_descanso_1, in_descanso_2, fim_descanso_2, " \
                f"in_descanso_3, fim_descanso_3, in_descanso_4, fim_descanso_4, in_descanso_5, fim_descanso_5, " \
                f"in_descanso_6, fim_descanso_6, in_descanso_7, fim_descanso_7, in_descanso_8, fim_descanso_8, " \
                f"in_car_desc_1, fim_car_desc_1, in_car_desc_2, fim_car_desc_2, in_car_desc_3, fim_car_desc_3, " \
                f"in_car_desc_4, fim_car_desc_4, in_car_desc_5, fim_car_desc_5, in_car_desc_6, fim_car_desc_6, " \
                f"in_car_desc_7, fim_car_desc_7 " \
                f"FROM perm_data " \
                f"JOIN motorists ON perm_data.motorist_id = motorists.id " \
                f"JOIN trucks ON perm_data.truck_id = trucks.id " \
                f"WHERE strftime('%Y-%m-%d', substr(data, 7, 4) || '-' || substr(data, 4, 2) || '-' || substr(data, 1, 2)) " \
                f"BETWEEN '{from_date}' AND '{to_date}' AND "

            arguments = []

            if motorist_id:
                arguments.append(f"motorist_id='{motorist_id}'")
            if truck_id:
                arguments.append(f"truck_id='{truck_id}'")

            arguments = " AND ".join(arguments)
            query = query + arguments

            conn = sqlite3.connect(DB_PATH)
            df_perm_data = pd.read_sql_query(query, conn)

            query = f"""
            SELECT data, hora, duration, desc_infracao 
            FROM infractions 
            WHERE strftime('%Y-%m-%d', substr(data, 7, 4) || '-' || substr(data, 4, 2) || '-' || substr(data, 1, 2)) 
            BETWEEN '{from_date}' AND '{to_date}' AND """
            query = query + arguments

            df_infractions = pd.read_sql_query(query, conn)

            # Criar o arquivo Excel se os dados existirem
            if not df_perm_data.empty:
                # Criar um arquivo Excel em memória
                output = BytesIO()
                wb = Workbook()
                ws = wb.active

                # Cabeçalhos para a planilha de dados de jornada
                journey_columns = [
                    "Nome Motorista", "Placa do Caminhão", "Data", "Dia da Semana", "Início da Jornada", "Fim Jornada",
                    "Observação", "Tempo de Refeição", "Tempo Intervalo", "Jornada Total", "Tempo de Direção", "Direção sem Pausa",
                    "1° Descanso Início", "1° Descanso Fim", "2° Descanso Início", "2° Descanso Fim", "3° Descanso Início",
                    "3° Descanso Fim", "4° Descanso Início", "4° Descanso Fim", "5° Descanso Início", "5° Descanso Fim",
                    "6° Descanso Início", "6° Descanso Fim", "7° Descanso Início", "7° Descanso Fim", "8° Descanso Início",
                    "8° Descanso Fim", "1° Carregamento Descanso Início", "1° Carregamento Descanso Fim",
                    "2° Carregamento Descanso Início", "2° Carregamento Descanso Fim", "3° Carregamento Descanso Início",
                    "3° Carregamento Descanso Fim", "4° Carregamento Descanso Início", "4° Carregamento Descanso Fim",
                    "5° Carregamento Descanso Início", "5° Carregamento Descanso Fim", "6° Carregamento Descanso Início",
                    "6° Carregamento Descanso Fim", "7° Carregamento Descanso Início", "7° Carregamento Descanso Fim"
                ]
                #

                # Cabeçalhos para a planilha de infrações
                infraction_columns = [
                    "Data", "Hora", "Duração", "Descrição"
                ]

                # Formatação do cabeçalho de dados de jornada
                header_fill = PatternFill(start_color="4176a8", end_color="4176a8", fill_type="solid")
                # Formatação do cabeçalho de infrações (fundo mais claro)
                infraction_header_fill = PatternFill(start_color="A9C8E3", end_color="A9C8E3", fill_type="solid")

                # Preencher os dados de jornada e infrações
                for i, row in df_perm_data.iterrows():
                    ws.append(journey_columns)

                    for cell in ws[ws.max_row]:
                        cell.fill = header_fill
                    # Adicionar dados de jornada
                    ws.append([row['nome'], row['placa'], row['data'], row['dia_da_semana'], row['inicio_jornada'], row['fim_jornada'],
                               row['observacao'], row['tempo_refeicao'], row['tempo_intervalo'], row['jornada_total'], row['tempo_direcao'],
                               row['direcao_sem_pausa'], row['in_descanso_1'], row['fim_descanso_1'], row['in_descanso_2'], row['fim_descanso_2'],
                               row['in_descanso_3'], row['fim_descanso_3'], row['in_descanso_4'], row['fim_descanso_4'], row['in_descanso_5'],
                               row['fim_descanso_5'], row['in_descanso_6'], row['fim_descanso_6'], row['in_descanso_7'], row['fim_descanso_7'],
                               row['in_descanso_8'], row['fim_descanso_8'], row['in_car_desc_1'], row['fim_car_desc_1'], row['in_car_desc_2'],
                               row['fim_car_desc_2'], row['in_car_desc_3'], row['fim_car_desc_3'], row['in_car_desc_4'], row['fim_car_desc_4'],
                               row['in_car_desc_5'], row['fim_car_desc_5'], row['in_car_desc_6'], row['fim_car_desc_6'], row['in_car_desc_7'],
                               row['fim_car_desc_7']])

                    # Adicionar uma linha em branco entre os dias
                    ws.append(['Infrações'])

                    # Adicionar as infrações do dia, se existirem
                    data_perm = row['data']
                    infractions_for_date = df_infractions[df_infractions['data'] == data_perm]

                    if not infractions_for_date.empty:
                        # Adicionar o cabeçalho de infrações
                        ws.append(infraction_columns)
                        # Aplicar o fundo mais claro para o cabeçalho de infrações
                        for cell in ws[ws.max_row]:
                            cell.fill = infraction_header_fill

                        # Preencher com as infrações desse dia
                        for _, infraction_row in infractions_for_date.iterrows():
                            ws.append([infraction_row['data'], infraction_row['hora'], infraction_row['duration'], infraction_row['desc_infracao']])

                    # Adicionar uma linha em branco após as infrações
                    ws.append([])

                # Ajustar a largura das colunas para que o conteúdo fique visível
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # A referência da coluna
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

                # Salvar o arquivo Excel em memória
                wb.save(output)
                output.seek(0)

                # Gerar o nome do arquivo com base nos parâmetros
                if motorist_id and truck_id:
                    motorist_name = df_perm_data.iloc[0]['nome']
                    plate = df_perm_data.iloc[0]['placa']
                    file_name = f"report_motorista_{motorist_name.replace(' ', '_')}_placa_{plate}_from_{from_date}_to_{to_date}.xlsx"
                elif motorist_id:
                    motorist_name = df_perm_data.iloc[0]['nome']
                    file_name = f"report_motorista_{motorist_name.replace(' ', '_')}_from_{from_date}_to_{to_date}.xlsx"
                elif truck_id:
                    plate = df_perm_data.iloc[0]['placa']
                    file_name = f"report_placa_{plate}_from_{from_date}_to_{to_date}.xlsx"

                # Retornar o arquivo Excel para o usuário
                return send_file(output, as_attachment=True,
                                 download_name=file_name,
                                 mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                flash("Nenhum registro encontrado.")
                return redirect("/download-report")
        except Exception as e:
            routes_logger.register_log(f"Ocorreu um erro durante a geração de relatório.", f"Erro: {e}")
            flash("Falha durante a geração de arquivo para download.")
            return redirect("/download-report")

    trucks = truck_driver.retrieve_all_trucks()
    motorists = motorist_driver.retrieve_all_motorists()
    # Ordenar motoristas alfabeticamente por nome
    motorists.sort(key=lambda x: x[1].upper() if x[1] else '')

    return render_template('track_reports.html', trucks=trucks, motorists=motorists)

@track_bp.route('/get_link_justification')
@route_access_required
def get_link_justification():
    try:
        infraction_hash = request.args.get('hash')
        if not infraction_hash:
            return 'Hash não fornecido', 400

        link = infraction_driver.get_link_justification(infraction_hash)
        return {'link': link} if link else {'link': None}
    except Exception as e:
        routes_logger.register_log("Erro ao buscar link da tratativa.", str(e))
        return 'Erro ao buscar link', 500
    
@track_bp.route('/update_link_justification', methods=['POST'])
@route_access_required
def update_link_justification():
    try:
        infraction_hash = request.form.get('hash')
        link = request.form.get('link')
        
        if not infraction_hash or not link:
            return 'Hash ou link não fornecido', 400

        # Atualiza o link da justification
        infraction_driver.update_link_justification(infraction_hash, link)
        
        return '', 204
    except Exception as e:
        routes_logger.register_log("Erro ao atualizar link da justification", str(e))
        return 'Erro ao atualizar link', 500
    

### API ROUTES

@track_bp.route('/api/get-infractions', methods=['POST'])
@route_access_required
def get_infractions():

    dados = request.get_json()
    tabela = dados.get('tabela')
    routes_logger.print(tabela)

    if isinstance(tabela, str):
        tabela = json.loads(tabela)

    motorist_id = dados.get('motorist_id', '')
    truck_id = dados.get('truck_id', '')

    try:
        df = convert_json_to_df(table_records=tabela,
                                motorist_id=motorist_id,
                                truck_id=truck_id)

        df['week_start'] = 0
        df['sorted_events'] = df.apply(get_sorted_events_with_work_periods, axis=1)
        infractions_list = compute_infractions(df)

        return_infractions = []

        for infraction in infractions_list:
            if infraction is not None:
                if infraction.get('infraction_type') is None:
                    continue
                else:
                    return_infractions.append(infraction)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"mensagem": "Erro ao gerar infrações", "erro": str(e)}), 400

    infractions_list = [infraction for infraction in infractions_list if infraction]
    infractions_list = [infraction for infraction in infractions_list if infraction.get('infraction_type')
                        in INFRACTION_DICT]

    return jsonify(infractions_list), 200

# Rota para pegar os N registros anteriores
@track_bp.route('/api/retrieve_n_records_before_date', methods=['GET'])
@route_access_required
def retrieve_n_records_before_date():
    motorist_id = request.args.get('motorist_id')
    truck_id = request.args.get('truck_id')
    date = request.args.get('target_date')
    n = int(request.args.get('n', 10))

    # Monta filtros
    where_columns = []
    where_values = []
    if motorist_id:
        where_columns.append('motorist_id')
        where_values.append(motorist_id)
    if truck_id:
        where_columns.append('truck_id')
        where_values.append(truck_id)

    records = perm_uploaded_track_driver.retrieve_n_records_before_date(
        n=n,
        target_date=date,
        where_columns=where_columns if where_columns else None,
        where_values=tuple(where_values) if where_values else (),
        output_format='dict'
    )
    return jsonify(records)

@track_bp.route('/api/save-table', methods=['POST'])
@route_access_required
def save_table():
    dados = request.get_json()

    tabela = dados.get('tabela')
    if isinstance(tabela, str):
        tabela = json.loads(tabela)

    motorist_id = dados.get('motorist_id')
    motorist_name = dados.get('motorist_name')
    truck_id = dados.get('truck_id')
    plate = dados.get('plate')
    acao = dados.get('acao', '')
    substituir = dados.get('substituir', False)  # Novo parâmetro para indicar se deve substituir
    
    # Novos campos para identificar registros novos
    datas_novas = dados.get('datas_novas', [])
    
    # Separar registros novos dos antigos
    registros_novos = []
    registros_antigos = []
    
    for registro in tabela:
        if registro.get('is_new_record', False) or registro.get('Data') in datas_novas:
            registros_novos.append(registro)
        else:
            registros_antigos.append(registro)
    
    # Para cálculo de infrações, usar todos os registros
    tabela_completa = tabela
    # Para salvamento, usar apenas os novos
    tabela_para_salvar = registros_novos

    if acao == "excel":
        excel_bytes = generate_excel(tabela)
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='relatorio.xlsx'
        )
    elif acao == "salvar":
        try:
            if substituir:
                # Substituir registros existentes (apenas os novos)
                result = perm_uploaded_track_driver.replace_data_from_json(tabela_para_salvar, motorist_id, truck_id)
                dayoff_result = dayoff_driver.replace_data_from_json(tabela_para_salvar, motorist_id)
            else:
                # NOVA LÓGICA: Verificar conflitos ANTES de inserir qualquer coisa
                # Usar apenas os registros novos para verificar conflitos
                result = perm_uploaded_track_driver.check_conflicts_only(tabela_para_salvar, motorist_id, truck_id)
                dayoff_result = dayoff_driver.check_conflicts_only(tabela_para_salvar, motorist_id)
                
                if result['tem_conflitos'] or dayoff_result['tem_conflitos']:
                    # Juntar os conflitos dos dois drivers
                    conflitos_detalhados = []
                    conflitos_simples = []
                    
                    # Processar conflitos de jornada (perm_data)
                    for conflito in result['conflitos']:
                        if isinstance(conflito, dict):
                            # Novo formato estruturado
                            conflitos_detalhados.append(conflito)
                            conflitos_simples.append(f"{conflito['data']} - {conflito['tipo']}")
                        else:
                            # Formato antigo para compatibilidade
                            conflitos_detalhados.append({
                                'data': conflito.split(' - ')[0] if ' - ' in conflito else conflito,
                                'tipo': 'Jornada',
                                'descricao': f"Já existe jornada registrada para {conflito}"
                            })
                            conflitos_simples.append(conflito)
                    
                    # Processar conflitos de dayoff
                    for conflito in dayoff_result['conflitos']:
                        if isinstance(conflito, dict):
                            # Novo formato estruturado
                            conflitos_detalhados.append(conflito)
                            conflitos_simples.append(f"{conflito['data']} - {conflito['tipo']}")
                        else:
                            # Formato antigo para compatibilidade
                            conflitos_detalhados.append({
                                'data': conflito.split(' - ')[0] if ' - ' in conflito else conflito,
                                'tipo': 'Dayoff',
                                'descricao': f"Já existe registro de folga para {conflito}"
                            })
                            conflitos_simples.append(conflito)
                    
                    # Retornar informações sobre conflitos para o frontend
                    return jsonify({
                        "mensagem": "Conflitos encontrados",
                        "status": "conflito",
                        "conflitos": list(set(conflitos_simples)),  # Para compatibilidade
                        "conflitos_detalhados": conflitos_detalhados
                    }), 409  # Código 409 para conflito
                
                # Se não há conflitos, inserir todos os registros novos de uma vez
                result = perm_uploaded_track_driver.insert_data_from_json(tabela_para_salvar, motorist_id, truck_id)
                dayoff_result = dayoff_driver.insert_data_from_json(tabela_para_salvar, motorist_id)
                
                routes_logger.print(f"Registros inseridos: {result['registros_inseridos']}, ignorados: {result['registros_ignorados']}")
                
            try:
                routes_logger.print(f"Apagando dados temporários da placa {plate}.")
                uploaded_track_driver.delete_record(where_columns=['truck_id', ],
                                          where_values=(truck_id,))
            except Exception as e:
                routes_logger.print(f"Erro ao excluir dados temporários. Erro: {e}.")

            # Salvando infrações (usar todos os registros para cálculo correto)
            try:
                df = convert_json_to_df(table_records=tabela_completa,
                                        motorist_id=motorist_id,
                                        truck_id=truck_id)

                if df is not None and not df.empty:
                    df['week_start'] = 0
                    df['sorted_events'] = df.apply(get_sorted_events_with_work_periods, axis=1)
                    routes_logger.print(f"Gerando infrações para o motorista {motorist_name}.")

                    try:
                        infractions_list = compute_infractions(df)
                    except:
                        traceback.print_exc()
                        raise

                    # Filtrar infrações apenas dos registros novos
                    datas_novas_set = set(datas_novas)
                    infracoes_novas = []
                    
                    for infraction in infractions_list:
                        if infraction is not None and infraction.get('infraction_type'):
                            # Verificar se a infração é de uma data nova
                            if infraction.get("date") in datas_novas_set:
                                infracoes_novas.append(infraction)
                    
                    # Salvar apenas as infrações dos registros novos
                    for infraction in infracoes_novas:
                        infraction_driver.create_infraction(motorist_id=motorist_id,
                                                            truck_id=truck_id,
                                                            data=infraction.get("date"),
                                                            hora=infraction.get("time"),
                                                            duration=infraction.get("duration"),
                                                            tipo_infracao=infraction.get('infraction_type'))

                    return jsonify({"mensagem": "Tabela recebida com sucesso!", "status": "ok"})

                else:
                    return jsonify({"mensagem": "Tabela recebida com sucesso mas não houve "
                                                "dias trabalhados para processar as infrações!", "status": "ok"})

            except Exception as e:
                routes_logger.register_log(f"Erro ao gerar infrações para o motorista {motorist_name}.", f'Erro: {e}')
                return jsonify({"mensagem": "Erro ao processar JSON", "erro": str(e)}), 400

        except Exception as e:
            routes_logger.print(f"Erro ao inserir dado permanente. Erro: {e}")
            return jsonify({"mensagem": "Erro ao processar dados", "erro": str(e)}), 500
    else:
        return jsonify({"mensagem": "Ação desconhecida."}), 400
    
@track_bp.route('/api/check_motorist_updates', methods=['GET'])
@route_access_required
def check_motorist_updates():
    try:
        # 1. Definir a Data-base como o dia anterior à data atual
        data_base = date.today() - timedelta(days=1)

        # 2. Recuperar apenas motoristas ativos para jornada e ordenar alfabeticamente
        all_motorists = motorist_driver.retrieve_active_motorists_for_journey()
        # Ordenar por nome (índice 1)
        all_motorists.sort(key=lambda x: x[1].upper() if x[1] else '')
        
        # 3. Extrair IDs dos motoristas para consultas em lote
        motorist_ids = [m[0] for m in all_motorists]
        
        # 4. Consultas em lote para otimizar performance
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Consulta em lote para todas as datas existentes (perm_data + dayoff)
        placeholders = ','.join(['?' for _ in motorist_ids])
        all_dates_query = f"""
            SELECT motorist_id, data FROM (
                SELECT motorist_id, data FROM perm_data WHERE motorist_id IN ({placeholders})
                UNION ALL
                SELECT motorist_id, data FROM dayoff WHERE motorist_id IN ({placeholders})
            )
        """
        cursor.execute(all_dates_query, motorist_ids + motorist_ids)
        all_dates_results = {}
        for row in cursor.fetchall():
            motorist_id = row[0]
            date_str = row[1]
            if motorist_id not in all_dates_results:
                all_dates_results[motorist_id] = set()
            all_dates_results[motorist_id].add(date_str)
        
        conn.close()
        
        update_status_list = []

        # 5. Processar cada motorista com os dados já carregados
        for motorist_info in all_motorists:
            motorist_id = motorist_info[0]
            motorist_name = motorist_info[1]
            motorist_admission_date = motorist_info[2]  # data_admissao (índice 2 na nova estrutura)

            # Definir data de início para verificação (data de admissão ou 01/05/2025, a mais recente)
            fixed_start_date = datetime.strptime('01/05/2025', '%d/%m/%Y').date()
            
            # Tentar converter a data de admissão
            admission_date = None
            if motorist_admission_date:
                try:
                    # Tentar diferentes formatos de data
                    for date_format in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y']:
                        try:
                            admission_date = datetime.strptime(motorist_admission_date, date_format).date()
                            break
                        except ValueError:
                            continue
                except:
                    pass
            
            # Usar a data mais recente entre admissão e 01/05/2025
            if admission_date:
                start_check_date = max(admission_date, fixed_start_date)
            else:
                start_check_date = fixed_start_date

            # Encontrar a última data de atualização
            existing_dates = all_dates_results.get(motorist_id, set())
            last_update_date = None
            last_update_display = "Sem registros"

            if existing_dates:
                # Converter todas as datas para objetos date para comparação
                date_objects = []
                for date_str in existing_dates:
                    try:
                        # Tentar diferentes formatos
                        for date_format in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d']:
                            try:
                                date_obj = datetime.strptime(date_str, date_format).date()
                                date_objects.append(date_obj)
                                break
                            except ValueError:
                                continue
                    except:
                        continue
                
                if date_objects:
                    last_update_date = max(date_objects)
                    last_update_display = last_update_date.strftime('%d-%m-%Y')

            # Calcular status baseado na última atualização
            status = "Sem registros"
            status_class = "sem-registros"

            if last_update_date:
                if last_update_date >= data_base:
                    status = "Atualizado"
                    status_class = "status-atualizado"
                else:
                    days_behind = (data_base - last_update_date).days
                    status = f"{days_behind} dias atrasado"
                    status_class = "status-atrasado"

            # Calcular observações (dias faltantes)
            obs = "Sem data de início para checagem"
            obs_class = "obs-faltante"

            if start_check_date and start_check_date <= data_base:
                missing_dates = []
                current_check_date = start_check_date
                
                while current_check_date <= data_base:
                    # Tentar diferentes formatos para verificar se a data existe
                    date_found = False
                    for date_format in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d']:
                        current_date_str = current_check_date.strftime(date_format)
                        if current_date_str in existing_dates:
                            date_found = True
                            break
                    
                    if not date_found:
                        missing_dates.append(current_check_date.strftime('%d/%m/%Y'))

                    current_check_date += timedelta(days=1)

                if not missing_dates:
                    obs = "Completo"
                    obs_class = "obs-completo"
                else:
                    obs = f"{', '.join(missing_dates)} faltante(s)"
                    obs_class = "obs-faltante"
            elif not start_check_date:
                obs = "Sem data de início para checagem"
                obs_class = "obs-faltante"
            else:
                obs = "Data de início futura"
                obs_class = ""

            update_status_list.append({
                "id": motorist_id,
                "name": motorist_name,
                "last_update": last_update_display,
                "status": status,
                "status_class": status_class,
                "obs": obs,
                "obs_class": obs_class
            })

        return jsonify(update_status_list), 200

    except Exception as e:
        routes_logger.register_log(f"Erro ao verificar atualizações de motoristas: {e}")
        return jsonify({"error": "Erro interno ao verificar atualizações."}), 500


@track_bp.route('/api/get-report', methods=['GET'])
@route_access_required
def get_report():
    motorist_id = request.args.get('motorist_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    if not from_date or not to_date:
        return "Datas obrigatórias não fornecidas", 400

    if isinstance(from_date, str):
        from_date = from_date.replace('/', '-')

    if isinstance(to_date, str):
        to_date = to_date.replace('/', '-')

    # Pegando informações de jornadas

    query = f"""
            SELECT motorists.nome, trucks.placa, data, dia_da_semana, inicio_jornada, in_refeicao, fim_refeicao, fim_jornada, 
            observacao, tempo_refeicao, intersticio, tempo_intervalo, tempo_carga_descarga, jornada_total, 
            tempo_direcao, direcao_sem_pausa 
            FROM perm_data 
            JOIN motorists ON perm_data.motorist_id = motorists.id 
            JOIN trucks ON perm_data.truck_id = trucks.id 
            WHERE strftime('%Y-%m-%d', substr(data, 7, 4) || '-' || substr(data, 4, 2) || '-' || substr(data, 1, 2)) 
            BETWEEN '{from_date}' AND '{to_date}' AND motorist_id='{motorist_id}'"""

    conn = sqlite3.connect(DB_PATH)
    df_perm_data = pd.read_sql_query(query, conn)

    conn.close()

    # Pegando informações de infrações

    query = f"""
                SELECT data, desc_infracao 
                FROM infractions 
                WHERE strftime('%Y-%m-%d', substr(data, 7, 4) || '-' || substr(data, 4, 2) || '-' || substr(data, 1, 2)) 
                BETWEEN '{from_date}' AND '{to_date}' AND motorist_id='{motorist_id}'"""

    conn = sqlite3.connect(DB_PATH)

    df_infractions = pd.read_sql_query(query, conn)

    conn.close()

    conn = sqlite3.connect(DB_PATH)

    # Pegando informações de folga...

    query = f"""
                    SELECT data, motivo 
                    FROM dayoff  
                    WHERE strftime('%Y-%m-%d', substr(data, 7, 4) || '-' || substr(data, 4, 2) || '-' || substr(data, 1, 2)) 
                    BETWEEN '{from_date}' AND '{to_date}' AND motorist_id='{motorist_id}'"""

    df_dayoff = pd.read_sql_query(query, conn)

    conn.close()

    # Montando o DF Final

    # Primeiro, garantir que as colunas 'data' estejam no formato datetime
    df_perm_data['data'] = pd.to_datetime(df_perm_data['data'], format="%d-%m-%Y")
    df_infractions['data'] = pd.to_datetime(df_infractions['data'], format="%d-%m-%Y")
    df_dayoff['data'] = pd.to_datetime(df_dayoff['data'], format="%d-%m-%Y")

    # Merge de infrações com df_perm_data
    df_infractions_grouped = df_infractions.groupby('data')['desc_infracao'] \
        .apply(lambda x: '\n'.join(f"- {item}" for item in x)).reset_index()
    df = pd.merge(df_perm_data, df_infractions_grouped, on='data', how='left')  # adiciona desc_infracao agrupado

    # Criando linhas extras para os dias de folga que não estão no df_perm_data
    existing_dates = set(df['data'])
    new_rows = []

    for _, row in df_dayoff.iterrows():
        if row['data'] not in existing_dates:
            new_row = {
                'nome': None,
                'placa': row['motivo'],  # motivo como placa
                'data': row['data'],
                'dia_da_semana': row['data'].weekday(),  # segunda=0, domingo=6
                'inicio_jornada': None,
                'in_refeicao': None,
                'fim_refeicao': None,
                'fim_jornada': None,
                'observacao': None,
                'tempo_refeicao': None,
                'intersticio': None,
                'tempo_intervalo': None,
                'tempo_carga_descarga': None,
                'jornada_total': None,
                'tempo_direcao': None,
                'direcao_sem_pausa': None,
                'desc_infracao': None
            }
            new_rows.append(new_row)

    # Cria DataFrame com as novas linhas
    df_new_dayoffs = pd.DataFrame(new_rows)

    # Juntando tudo em um DF
    df_final = pd.concat([df, df_new_dayoffs], ignore_index=True)

    # Ordena por data
    df_final = df_final.sort_values(by='data').reset_index(drop=True)

    # Mapeamento de números para nomes dos dias da semana
    dias_semana = {
        0: 'Seg.',
        1: 'Ter.',
        2: 'Qua.',
        3: 'Qui.',
        4: 'Sex.',
        5: 'Sáb.',
        6: 'Dom.'
    }

    df_final['dia_da_semana'] = df_final['data'].dt.weekday.map(dias_semana)
    df_final = df_final.fillna('')

    df_dict = df_final.to_dict(orient='records')

    motorist = motorist_driver.retrieve_motorist(where_columns=['id', ], where_values=(motorist_id,))
    motorist_name = motorist[1]

    data_obj = datetime.strptime(from_date, "%Y-%m-%d")
    from_date = data_obj.strftime("%d/%m/%Y")
    data_obj = datetime.strptime(to_date, "%Y-%m-%d")
    to_date = data_obj.strftime("%d/%m/%Y")

    return render_template('partials/result_modal_partial.html', resultados=df_dict,
                           nome=motorist_name,
                           inicio=from_date,
                           fim=to_date)


@track_bp.route('/api/export_excel', methods=['POST'])
@route_access_required
def export_excel():
    data = request.json
    if not data or not data.get('nome_motorista') or not data.get('data_inicio') or not data.get('data_fim') or not data.get('tabela') or not data.get('totais'):
        return jsonify({"error": "Dados incompletos para exportar Excel"}), 400

    wb = fill_excel(
        name=data['nome_motorista'],
        start=data['data_inicio'],
        end=data['data_fim'],
        tabela=data['tabela'],
        totals=data['totais']
    )
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"jornada_motorista_{data['nome_motorista'].replace(' ', '_')}_" \
               f"{data['data_inicio'].replace('/', '-')}_a_{data['data_fim'].replace('/', '-')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@track_bp.route('/api/export_pdf', methods=['POST'])
@route_access_required
def export_pdf():
    """
    Endpoint para gerar PDF usando reportlab (substitui o método antigo do Excel).
    """
    data = request.json
    
    if not data or not data.get('nome_motorista') or not data.get('data_inicio') or not data.get('data_fim') or not data.get('tabela') or not data.get('totais'):
        return jsonify({"error": "Dados incompletos para exportar PDF"}), 400

    try:
        # Gerar PDF usando reportlab
        pdf_bytes = fill_pdf(
            name=data['nome_motorista'],
            start=data['data_inicio'],
            end=data['data_fim'],
            tabela=data['tabela'],
            totals=data['totais']
        )
        
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'jornada_motorista_{data["nome_motorista"].replace(" ", "_")}_{data["data_inicio"].replace("/", "-")}_a_{data["data_fim"].replace("/", "-")}.pdf'
        )
        
    except Exception as e:
        routes_logger.register_log(f"Erro ao gerar PDF com reportlab: {e}")
        return jsonify({'error': 'Erro ao gerar PDF'}), 500


@track_bp.route('/api/get-motorist-details', methods=['GET'])
@route_access_required
def get_motorist_details():
    motorist_id = request.args.get('motorist_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    if not all([motorist_id, from_date, to_date]):
        return jsonify({"error": "Parâmetros obrigatórios não fornecidos"}), 400

    try:
        # Consulta para pegar dados de jornada e folgas
        query = """
            SELECT 
                COALESCE(p.data, d.data) as data,
                CASE 
                    WHEN d.data IS NOT NULL THEN d.motivo
                    ELSE t.placa 
                END as placa,
                CASE 
                    WHEN d.data IS NOT NULL THEN NULL
                    ELSE p.dia_da_semana 
                END as dia_da_semana,
                p.inicio_jornada,
                p.in_refeicao,
                p.fim_refeicao,
                p.fim_jornada,
                p.observacao,
                p.tempo_refeicao,
                p.intersticio,
                p.tempo_intervalo,
                p.tempo_carga_descarga,
                p.jornada_total,
                p.tempo_direcao,
                p.direcao_sem_pausa
            FROM 
                (SELECT data FROM perm_data WHERE motorist_id = ? 
                 UNION 
                 SELECT data FROM dayoff WHERE motorist_id = ?) dates
            LEFT JOIN perm_data p ON dates.data = p.data AND p.motorist_id = ?
            LEFT JOIN dayoff d ON dates.data = d.data AND d.motorist_id = ?
            LEFT JOIN trucks t ON p.truck_id = t.id
            WHERE strftime('%Y-%m-%d', substr(dates.data, 7, 4) || '-' || substr(dates.data, 4, 2) || '-' || substr(dates.data, 1, 2)) 
            BETWEEN ? AND ?
            ORDER BY strftime('%Y-%m-%d', substr(dates.data, 7, 4) || '-' || substr(dates.data, 4, 2) || '-' || substr(dates.data, 1, 2)) DESC
        """

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute(query, (motorist_id, motorist_id, motorist_id, motorist_id, from_date, to_date))
        rows = cursor.fetchall()
        
        # Consulta para pegar infrações
        infraction_query = """
            SELECT data, GROUP_CONCAT(desc_infracao, '\n') as infractions
            FROM infractions 
            WHERE motorist_id = ? AND
            strftime('%Y-%m-%d', substr(data, 7, 4) || '-' || substr(data, 4, 2) || '-' || substr(data, 1, 2)) 
            BETWEEN ? AND ?
            GROUP BY data
        """
        
        cursor.execute(infraction_query, (motorist_id, from_date, to_date))
        infractions = {row[0]: row[1] for row in cursor.fetchall()}
        
        conn.close()

        # Converter resultados para dicionário
        results = []
        columns = ['data', 'placa', 'dia_da_semana', 'inicio_jornada', 'in_refeicao', 
                  'fim_refeicao', 'fim_jornada', 'observacao', 'tempo_refeicao', 
                  'intersticio', 'tempo_intervalo', 'tempo_carga_descarga', 
                  'jornada_total', 'tempo_direcao', 'direcao_sem_pausa']
        
        for row in rows:
            result = dict(zip(columns, row))
            result['desc_infracao'] = infractions.get(result['data'], '')
            results.append(result)

        return jsonify(results)

    except Exception as e:
        routes_logger.register_log(f"Erro ao buscar detalhes do motorista: {str(e)}")
        return jsonify({"error": "Erro interno ao buscar detalhes"}), 500


@track_bp.route('/api/delete-motorist-records', methods=['POST'])
@route_access_required
def delete_motorist_records():
    data = request.get_json()
    motorist_id = data.get('motorist_id')
    dates = data.get('dates', [])

    if not motorist_id or not dates:
        return jsonify({"error": "Parâmetros obrigatórios não fornecidos"}), 400

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Excluir registros de jornada
        placeholders = ','.join(['?' for _ in dates])
        cursor.execute(f"""
            DELETE FROM perm_data 
            WHERE motorist_id = ? AND data IN ({placeholders})
        """, [motorist_id] + dates)

        # Excluir registros de folga
        cursor.execute(f"""
            DELETE FROM dayoff 
            WHERE motorist_id = ? AND data IN ({placeholders})
        """, [motorist_id] + dates)

        # Excluir infrações associadas
        cursor.execute(f"""
            DELETE FROM infractions 
            WHERE motorist_id = ? AND data IN ({placeholders})
        """, [motorist_id] + dates)

        conn.commit()
        conn.close()

        return jsonify({"message": "Registros excluídos com sucesso"})

    except Exception as e:
        routes_logger.register_log(f"Erro ao excluir registros do motorista: {str(e)}")
        return jsonify({"error": "Erro interno ao excluir registros"}), 500


@track_bp.route('/api/substituir-registros', methods=['POST'])
@route_access_required
def substituir_registros():
    """
    Rota para substituir registros existentes quando há conflitos.
    Implementa lógica "tudo ou nada" - substitui todos os registros de uma vez.
    """
    dados = request.get_json()

    tabela = dados.get('tabela')
    if isinstance(tabela, str):
        tabela = json.loads(tabela)

    motorist_id = dados.get('motorist_id')
    motorist_name = dados.get('motorist_name')
    truck_id = dados.get('truck_id')
    plate = dados.get('plate')

    if not all([tabela, motorist_id, truck_id]):
        return jsonify({"mensagem": "Dados incompletos para substituição"}), 400

    try:
        # Extrair datas da tabela para remover infrações antigas
        datas_para_substituir = []
        for linha in tabela:
            if linha.get('Data'):
                datas_para_substituir.append(linha['Data'])
        
        # Remover infrações antigas associadas às datas que serão substituídas
        if datas_para_substituir:
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                
                # Criar placeholders para a query IN
                placeholders = ','.join(['?' for _ in datas_para_substituir])
                
                # Remover infrações antigas
                cursor.execute(f"""
                    DELETE FROM infractions 
                    WHERE motorist_id = ? AND data IN ({placeholders})
                """, [motorist_id] + datas_para_substituir)
                
                infracoes_removidas = cursor.rowcount
                routes_logger.print(f"Infrações antigas removidas: {infracoes_removidas}")
                
                conn.commit()
                conn.close()
            except Exception as e:
                routes_logger.print(f"Erro ao remover infrações antigas: {e}")
        
        # Substituir registros existentes (todos de uma vez)
        result = perm_uploaded_track_driver.replace_data_from_json(tabela, motorist_id, truck_id)
        routes_logger.print(f"Registros substituídos: {result['registros_substituidos']}")

        # Processar folgas (todos de uma vez)
        try:
            dayoff_result = dayoff_driver.replace_data_from_json(tabela, motorist_id)
            routes_logger.print(f"Folgas substituídas: {dayoff_result['registros_substituidos']}")
        except Exception as e:
            routes_logger.print(f"Erro ao inserir dado de folga. Erro: {e}")

        # Apagar dados temporários
        try:
            routes_logger.print(f"Apagando dados temporários da placa {plate}.")
            uploaded_track_driver.delete_record(where_columns=['truck_id'], where_values=(truck_id,))
        except Exception as e:
            routes_logger.print(f"Erro ao excluir dados temporários. Erro: {e}.")

        # Processar infrações (usar todos os registros para cálculo correto)
        try:
            df = convert_json_to_df(table_records=tabela,
                                    motorist_id=motorist_id,
                                    truck_id=truck_id)

            if df is not None and not df.empty:
                df['week_start'] = 0
                df['sorted_events'] = df.apply(get_sorted_events_with_work_periods, axis=1)
                routes_logger.print(f"Gerando infrações para o motorista {motorist_name}.")

                infractions_list = compute_infractions(df)
                routes_logger.register_log(f"Infrações registradas no banco de dados: {infractions_list}")

                for infraction in infractions_list:
                    if infraction is not None and infraction.get('infraction_type'):
                        infraction_driver.create_infraction(motorist_id=motorist_id,
                                                            truck_id=truck_id,
                                                            data=infraction.get("date"),
                                                            hora=infraction.get("time"),
                                                            duration=infraction.get("duration"),
                                                            tipo_infracao=infraction.get('infraction_type'))

        except Exception as e:
            routes_logger.register_log(f"Erro ao gerar infrações para o motorista {motorist_name}.", f'Erro: {e}')

        return jsonify({
            "mensagem": "Registros substituídos com sucesso!", 
            "status": "ok",
            "registros_substituidos": result['registros_substituidos']
        })

    except Exception as e:
        routes_logger.print(f"Erro ao substituir registros. Erro: {e}")
        return jsonify({"mensagem": "Erro ao substituir registros", "erro": str(e)}), 500