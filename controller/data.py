import traceback

try:
    import pandas as pd
except ImportError:
    print("AVISO: pandas não disponível, usando stub")
    import pandas_stub as pd
from datetime import datetime, timedelta
from controller.utils import seconds_to_str_HM, convert_date_format
from model.drivers.truck_driver import TruckDriver
from global_vars import DB_PATH
import sqlite3
import xlrd

from typing import List, Dict

from global_vars import ALLOWED_EXTENSIONS, DEBUG

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side
from copy import copy
import io
import tempfile
import os
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import math
import re
from controller.utils import CustomLogger


def convert_data(data_str: str, mode="to_iso") -> str:
    """
    Converte uma data no formato 'dd:mm:YYYY' para 'YYYY:mm:dd' ou vice-versa.

    Parâmetros:
        data_str (str): Data no formato 'dd:mm:YYYY' ou 'YYYY:mm:dd'.
        mode (str): Define a conversão. Se "to_iso", converte para 'YYYY:mm:dd'. Caso contrário,
        converte para 'dd:mm:YYYY'.

    Retorna:
        str: Data convertida no formato desejado.
    """
    if mode == "to_iso":
        dia, mes, ano = data_str.split('-')
        return f"{ano}-{mes}-{dia}"
    else:
        ano, mes, dia = data_str.split('-')
        return f"{dia}-{mes}-{ano}"


def allowed_file(filename: str) -> bool:
    """
    Verifica se o arquivo possui uma extensão permitida.

    Parâmetros:
        filename (str): Nome do arquivo.

    Retorna:
        bool: True se a extensão do arquivo for permitida, caso contrário, False.
    """
    return '.' in filename and filename.rsplit('.')[-1].lower() in ALLOWED_EXTENSIONS


def make_data_block(records_df, date) -> List[Dict] | dict:
    """
    Cria um bloco de dados com base nos registros fornecidos e na data.

    Parâmetros:
        records_df (pd.DataFrame): DataFrame com os registros de dados de rastreamento.
        date (str): Data no formato 'YYYY-MM-DD'.

    Retorna:
        dict: Dicionário contendo o bloco de dados com informações de jornada, paradas, e se há movimento ou não.
    """
    week_days = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira',
                 'Sexta-feira', 'Sábado', 'Domingo']
    date_dt = datetime.strptime(date, '%Y-%m-%d')

    week_day = week_days[date_dt.weekday()]

    df_records_by_speed, \
    first_work_start_speed, \
    last_work_end_speed, \
    first_work_coords_speed, \
    last_work_coords_speed = generate_rests_df(records_df, mode='vel')

    df_records_by_ign, \
    first_work_start_ign, \
    last_work_end_ign, \
    first_work_coords_ign, \
    last_work_coords_ign = generate_rests_df(records_df, mode='ignicao')

    formatted_date = convert_date_format(date)

    bloco = {
        "data_bloco": formatted_date,
        "dia_semana": week_day,
        "inicio_movimento": None,
        "inicio_ignicao": None,
        "fim_movimento": None,
        "fim_ignicao": None,
        "paradas": [],
        "sem_movimento": False,
        "local_sem_movimento": None
    }

    if df_records_by_speed['type'].nunique() == 1 and df_records_by_speed['type'].iloc[0] == 'rest':
        bloco["sem_movimento"] = True
        row = df_records_by_speed.iloc[0]
        bloco["local_sem_movimento"] = {
            "lat": row['latitude'],
            "lon": row['longitude'],
            "link": f"https://www.google.com/maps?q={row['latitude']}, {row['longitude']}"
        }
        if df_records_by_ign['type'].nunique() == 2:  # Não tem registro de velocidade mas tem de ignição
            bloco["inicio_ignicao"] = {
                "hora": first_work_start_ign,
                "lat": first_work_coords_ign[0],
                "lon": first_work_coords_ign[1],
                "link": f"https://www.google.com/maps?q={first_work_coords_ign[0]}, {first_work_coords_ign[1]}"
            }
            bloco["fim_ignicao"] = {
                "hora": last_work_end_ign,
                "lat": last_work_coords_ign[0],
                "lon": last_work_coords_ign[1],
                "link": f"https://www.google.com/maps?q={last_work_coords_ign[0]}, {last_work_coords_ign[1]}"
            }

        return bloco

    try:
        bloco["inicio_movimento"] = {
            "hora": first_work_start_speed,
            "lat": first_work_coords_speed[0],
            "lon": first_work_coords_speed[1],
            "link": f"https://www.google.com/maps?q={first_work_coords_speed[0]}, {first_work_coords_speed[1]}"
        }

        bloco["inicio_ignicao"] = {
            "hora": first_work_start_ign,
            "lat": first_work_coords_ign[0],
            "lon": first_work_coords_ign[1],
            "link": f"https://www.google.com/maps?q={first_work_coords_ign[0]}, {first_work_coords_ign[1]}"
        }

        bloco["fim_movimento"] = {
            "hora": last_work_end_speed,
            "lat": last_work_coords_speed[0],
            "lon": last_work_coords_speed[1],
            "link": f"https://www.google.com/maps?q={last_work_coords_speed[0]}, {last_work_coords_speed[1]}"
        }

        bloco["fim_ignicao"] = {
            "hora": last_work_end_ign,
            "lat": last_work_coords_ign[0],
            "lon": last_work_coords_ign[1],
            "link": f"https://www.google.com/maps?q={last_work_coords_ign[0]}, {last_work_coords_ign[1]}"
        }
    except Exception:
        traceback.print_exc()
        raise

    for _, row in df_records_by_speed.iterrows():
        if row['type'] == 'rest':
            bloco["paradas"].append({
                "inicio": row['start'].strftime('%H:%M'),
                "fim": row['end'].strftime('%H:%M'),
                "lat": row['latitude'],
                "lon": row['longitude'],
                "cidade": row['cidade'],
                "rua": row['rua'],
                "link": f"https://www.google.com/maps?q={row['latitude']},{row['longitude']}",
                "tempo": seconds_to_str_HM(row['duration']),
                "validacao": "vel"
            })
    return bloco




def fill_work_periods(segments: pd.DataFrame):
    """
    Preenche os períodos de trabalho entre segmentos de descanso.

    Parâmetros:
        segments (pd.DataFrame): DataFrame contendo segmentos de descanso e trabalho.

    Retorna:
        pd.DataFrame: DataFrame atualizado com os períodos de trabalho preenchidos.
    """
    segments = segments.sort_values('start').reset_index(drop=True)
    filled_segments = []

    for i in range(len(segments)):
        # Converte linha para dicionário antes de adicionar
        filled_segments.append(segments.iloc[i].to_dict())

        if i < len(segments) - 1:
            end_current = segments.iloc[i]['end']
            start_next = segments.iloc[i + 1]['start']

            if (start_next - end_current).total_seconds() > 0:
                # Adiciona período de trabalho entre descansos
                filled_segments.append({
                    'start': end_current,
                    'end': start_next,
                    'duration': (start_next - end_current).total_seconds(),
                    'type': 'work',
                    'latitude': None,
                    'longitude': None,
                    'street': None,
                    'city': None,
                    'state': None
                })

    return pd.DataFrame(filled_segments)


def split_by_day(start, end, row_info):
    """
    Divide um período entre duas datas em segmentos diários.

    Parâmetros:
        start (datetime): Data/hora de início do período.
        end (datetime): Data/hora de fim do período.
        row_info (dict): Informações adicionais a serem incluídas em cada segmento diário.

    Retorna:
        list: Lista de segmentos diários, com cada segmento representado como um dicionário.
    """
    current = start
    segments = []

    while current.date() < end.date():
        next_midnight = datetime.combine(current.date() + pd.Timedelta(days=1), datetime.min.time())
        duration = (next_midnight - current).total_seconds()

        if duration > 0:
            segments.append({
                'start': current,
                'end': next_midnight,
                'duration_seconds': duration,
                **row_info
            })
        current = next_midnight

    duration = (end - current).total_seconds()
    if duration > 0:
        segments.append({
            'start': current,
            'end': end,
            'duration_seconds': duration,
            **row_info
        })

    return segments

def generate_rests_df(df: pd.DataFrame, mode='vel', initial_stat=None, include_prepost_rest=False):
    """
    Gera um DataFrame de segmentos de descanso e trabalho com base na velocidade ou ignição.

    Parâmetros:
        df (pd.DataFrame): DataFrame de entrada com colunas obrigatórias:
            - data_iso, vel ou ignicao, latitude, longitude, cidade, rua
        mode (str): 'vel' ou 'ignicao'
        initial_stat (str | None): Estado inicial ('rest' ou 'work'). Se None, é inferido da primeira linha.
        include_prepost_rest (bool): Se False, remove segmentos de descanso antes do primeiro 'work'
                                     e após o último 'work'.

    Retorna:
        Tuple[
            pd.DataFrame,        # Segmentos (rest/work)
            pd.Timestamp | None, # Início da jornada (primeiro work)
            pd.Timestamp | None, # Fim da jornada (último work)
            Tuple[float, float] | None,  # Coordenadas de início do trabalho
            Tuple[float, float] | None   # Coordenadas de fim do trabalho
        ]
    """
    assert mode in ['vel', 'ignicao'], "Modo deve ser 'vel' ou 'ignicao'"

    # Função interna que verifica o status de descanso ou trabalho baseado em dados da linha
    def get_status(row, prev_row=None):
        """
        Retorna o estado atual (descanso ou trabalho) com base na velocidade e nas coordenadas.

        - Se o modo for 'vel', considera a velocidade e se a latitude/longitude mudaram.
        - Se o modo for 'ignicao', considera a ignição e se a latitude/longitude mudaram.
        """
        if mode == 'vel':
            # Se não tem linha anterior, assume que tá descansando
            if prev_row is None:
                return 'rest'

            # Verifica se tá rodando (velocidade > 0, coordenadas mudaram e ignição ligada)
            if row['vel'] != 0 and row['latitude'] != prev_row['latitude'] and \
                    row['longitude'] != prev_row['longitude'] and row['ignicao'] == 'Ligada':
                return 'work'
            else:
                return 'rest'  # Se não tá rodando, tá descansando
        else:
            # No modo ignição, só olha se tá ligada ou desligada
            if row['ignicao'] == 'Desligada':
                return 'rest'
            else:
                return 'work'

    df['data_iso'] = pd.to_datetime(df['data_iso'])
    df = df.sort_values(by='data_iso').reset_index(drop=True)

    threshold = timedelta(minutes=5)
    all_segments = []
    buffer = []

    if not initial_stat:
        current_stat = get_status(df.iloc[0], prev_row=None)
    else:
        current_stat = initial_stat

    start_time = df.iloc[0]['data_iso']
    start_coords = (df.iloc[0]['latitude'], df.iloc[0]['longitude'])

    first_work_start = None
    last_work_end = None
    first_work_coords = None
    last_work_coords = None

    for i in range(1, len(df)):
        row = df.iloc[i]
        prev_row = df.iloc[i - 1]

        new_stat = get_status(row, prev_row=prev_row)

        if new_stat == current_stat:
            buffer.clear()
            continue

        if mode == 'vel':
            if current_stat == 'work' and new_stat == 'rest':
                buffer.append(row)
                buffer_duration = buffer[-1]['data_iso'] - buffer[0]['data_iso']

                # Só considera descanso se durar pelo menos 5 minutos
                if buffer_duration >= threshold:
                    prev_row = df.iloc[i - len(buffer)]
                    segment_start = start_time
                    segment_end = buffer[0]['data_iso']
                    segment_duration = (segment_end - segment_start).total_seconds()

                    all_segments.append({
                        'start': segment_start,
                        'end': segment_end,
                        'duration': segment_duration,
                        'type': current_stat,
                        'latitude': start_coords[0],
                        'longitude': start_coords[1],
                        'end_latitude': prev_row['latitude'],
                        'end_longitude': prev_row['longitude'],
                        'cidade': prev_row['cidade'],
                        'rua': prev_row['rua']
                    })

                    # Guarda info do primeiro e último trabalho
                    if not first_work_start:
                        first_work_start = segment_start
                        first_work_coords = start_coords
                    last_work_end = segment_end
                    last_work_coords = (prev_row['latitude'], prev_row['longitude'])

                    current_stat = new_stat
                    start_time = buffer[0]['data_iso']
                    start_coords = (buffer[0]['latitude'], buffer[0]['longitude'])
                    buffer.clear()

            elif current_stat == 'rest' and new_stat == 'work':
                # Quando muda de descanso pra trabalho, troca na hora
                segment_start = start_time
                segment_end = row['data_iso']
                segment_duration = (segment_end - segment_start).total_seconds()

                all_segments.append({
                    'start': segment_start,
                    'end': segment_end,
                    'duration': segment_duration,
                    'type': current_stat,
                    'latitude': start_coords[0],
                    'longitude': start_coords[1],
                    'end_latitude': row['latitude'],
                    'end_longitude': row['longitude'],
                    'cidade': row['cidade'],
                    'rua': row['rua']
                })

                current_stat = new_stat
                start_time = row['data_iso']
                start_coords = (row['latitude'], row['longitude'])
                buffer.clear()

        elif mode == 'ignicao':
            if current_stat == 'work' and new_stat == 'rest':
                # No modo ignição, troca na hora quando desliga
                segment_start = start_time
                segment_end = row['data_iso']
                segment_duration = (segment_end - segment_start).total_seconds()

                all_segments.append({
                    'start': segment_start,
                    'end': segment_end,
                    'duration': segment_duration,
                    'type': current_stat,
                    'latitude': start_coords[0],
                    'longitude': start_coords[1],
                    'end_latitude': row['latitude'],
                    'end_longitude': row['longitude'],
                    'cidade': row['cidade'],
                    'rua': row['rua']
                })

                # Guarda info do primeiro e último trabalho
                if not first_work_start:
                    first_work_start = segment_start
                    first_work_coords = start_coords
                last_work_end = segment_end
                last_work_coords = (prev_row['latitude'], prev_row['longitude'])

                current_stat = new_stat
                start_time = row['data_iso']
                start_coords = (row['latitude'], row['longitude'])
                buffer.clear()

            elif current_stat == 'rest' and new_stat == 'work':
                # Quando liga a ignição, começa a trabalhar
                segment_start = start_time
                segment_end = row['data_iso']
                segment_duration = (segment_end - segment_start).total_seconds()

                all_segments.append({
                    'start': segment_start,
                    'end': segment_end,
                    'duration': segment_duration,
                    'type': current_stat,
                    'latitude': start_coords[0],
                    'longitude': start_coords[1],
                    'end_latitude': row['latitude'],
                    'end_longitude': row['longitude'],
                    'cidade': row['cidade'],
                    'rua': row['rua']
                })

                current_stat = new_stat
                start_time = row['data_iso']
                start_coords = (row['latitude'], row['longitude'])
                buffer.clear()

    # Processa o último segmento
    if len(df) > 0:
        last_row = df.iloc[-1]
        last_ts = last_row['data_iso']

        # Se ainda tá trabalhando no final, termina o segmento na última linha
        if current_stat == 'work':
            segment_end = last_ts
        else:
            segment_end = last_ts

        segment_duration = (segment_end - start_time).total_seconds()

        all_segments.append({
            'start': start_time,
            'end': segment_end,
            'duration': segment_duration,
            'type': current_stat,
            'latitude': start_coords[0],
            'longitude': start_coords[1],
            'end_latitude': last_row['latitude'],
            'end_longitude': last_row['longitude'],
            'cidade': last_row['cidade'],
            'rua': last_row['rua']
        })

        # Atualiza as coordenadas do último trabalho se necessário
        if current_stat == 'work':
            if not first_work_start:
                first_work_start = start_time
                first_work_coords = start_coords
            last_work_end = segment_end
            last_work_coords = (last_row['latitude'], last_row['longitude'])

    segments_df = pd.DataFrame(all_segments)

    # Remove descansos antes do primeiro trabalho e depois do último (se não quiser incluir)
    if not include_prepost_rest:
        if first_work_start and last_work_end:
            segments_df = segments_df[
                ~((segments_df['type'] == 'rest') &
                  ((segments_df['end'] <= first_work_start) |
                   (segments_df['start'] >= last_work_end)))
            ].reset_index(drop=True)

    return segments_df, first_work_start, last_work_end, first_work_coords, last_work_coords


def extract_data(filepath: str, system_type: str, truck_driver: TruckDriver = None, 
                 logger: CustomLogger = None, auto_create_trucks: bool = True) -> pd.DataFrame:
    """
    Extrai e processa os dados de um arquivo de rastreamento, convertendo-os em um DataFrame do Pandas.

    Esta função lê um arquivo de rastreamento (Excel ou CSV), dependendo do tipo de sistema de
    rastreamento especificado.

    A função realiza as seguintes operações:
    - Converte as colunas de data e hora para o formato ISO.
    - Converte a coluna de velocidade para um número inteiro.
    - Converte as coordenadas de latitude e longitude.
    - Trata a coluna de ignição (ligada/desligada).
    - Atribui IDs de caminhão ao DataFrame, verificando se as placas são válidas no banco de dados.

    Parâmetros:
        filepath (str): Caminho para o arquivo a ser lido. Pode ser um arquivo Excel (.xlsx) ou CSV (.csv).
        system_type (str): Tipo do sistema de rastreamento. Deve ser um dos seguintes: 'sasgc', 'sascar', 'positron'.
        truck_driver (TruckDriver, opcional): Objeto opcional para manipulação do banco de dados de caminhões.
        logger (CustomLogger, opcional): Logger personalizado. Se não fornecido, cria um padrão.
        auto_create_trucks (bool): Se True, cria automaticamente caminhões não encontrados no banco. Default: True.

    Retorna:
        pd.DataFrame: DataFrame contendo os dados extraídos e processados. As colunas incluem:
            - truck_id (int): ID do caminhão.
            - data_iso (str): Data e hora no formato ISO.
            - vel (int): Velocidade do caminhão.
            - latitude (float): Latitude da posição.
            - longitude (float): Longitude da posição.
            - uf (str): Unidade federativa (estado).
            - cidade (str): Cidade.
            - rua (str): Rua.
            - ignicao (str): Estado da ignição ('Ligada' ou 'Desligada').

    Levanta:
        AttributeError: Se o system_type fornecido não for um dos tipos permitidos ('sasgc', 'sascar', 'positron').
    """
    
    # Inicializa logger se não fornecido
    if logger is None:
        logger = CustomLogger(source="DATA_EXTRACTION", debug=DEBUG)
    
    logger.register_log(f'[DEBUG] Início do extract_data para arquivo: {filepath}, tipo: {system_type}')

    # Normaliza o tipo do sistema de rastreamento para minúsculas
    system_type = system_type.lower()

    # Lista de tipos de sistemas de rastreamento permitidos
    allowed_types = ['sasgc', 'sascar', 'positron']

    # Verifica se o tipo de sistema é permitido
    if system_type not in allowed_types:
        logger.register_log(f'[ERRO] Tipo de rastreador não permitido: {system_type}')
        raise AttributeError(f"Tipo não permitido. O Atributo 'system_type' deve ser entre "
                             f"esses valores: {allowed_types}. Valor fornecido: {system_type}")

    # Define as colunas de saída para o DataFrame
    output_columns = ["truck_id", "data_iso", "vel", "latitude", "longitude", "uf", "cidade", "rua", "ignicao"]
    output_df = pd.DataFrame(columns=output_columns)

    # Função para corrigir a data de arquivos Positron no formato XLSX
    def fix_positron_date(date_str):
        """
        Corrige a data de arquivos Positron no formato XLSX para o padrão ISO.
        
        Parâmetros:
            date_str (str | datetime): Data no formato string ou datetime.

        Retorna:
            str: Data formatada no padrão ISO (YYYY-MM-DD HH:MM:SS) ou None se não for possível converter.
        """
        # Verifica se já é um datetime
        if isinstance(date_str, datetime):
            return date_str.strftime('%Y-%m-%d %H:%M:%S')
        
        # Se não for datetime, tenta converter manualmente
        if isinstance(date_str, str):
            formatos_tentativa = [
                '%d/%m/%Y %H:%M:%S',   # Dia/Mês/Ano
                '%m/%d/%Y %H:%M:%S',   # Mês/Dia/Ano
                '%Y/%m/%d %H:%M:%S',   # Ano/Mês/Dia
                '%Y-%m-%d %H:%M:%S'    # Ano-Mês-Dia
            ]
            
            for formato in formatos_tentativa:
                try:
                    data = datetime.strptime(date_str, formato)
                    # Se conseguiu converter, retorna no formato ISO
                    return data.strftime('%Y-%m-%d %H:%M:%S')
                except ValueError:
                    continue

        # Se todas as tentativas falharem, retorna None
        return None

    # Caso o sistema de rastreamento seja 'positron'
    if system_type == 'positron':
        # 🔍 Detecta a extensão do arquivo para escolher o engine correto
        file_extension = filepath.split('.')[-1].lower()

        # ✅ Leitura de arquivo Excel com o engine apropriado
        if file_extension == 'xls':
            df = pd.read_excel(filepath)
        elif file_extension == 'xlsx':
            df = pd.read_excel(filepath, engine='openpyxl', dtype=str)  # <=== Forçando string
        else:
            logger.register_log(f'[ERRO] Formato de arquivo não suportado para POSITRON: {file_extension}')
            raise ValueError(f"Formato de arquivo não suportado para POSITRON: {file_extension}")

        # Ajusta o cabeçalho do DataFrame
        df.columns = df.iloc[0]
        df = df[1:]  # Remove a primeira linha que virou cabeçalho
        df = df.reset_index(drop=True)  # Reseta os índices

        # Cria o DataFrame de saída
        output_df = pd.DataFrame()

        # ✅ Converte a coluna de Horário para string
        df["Horário"] = df["Horário"].astype(str)

        # ✅ Aplica a correção de datas
        df["Horário"] = df["Horário"].apply(fix_positron_date)

        # ✅ Salvando no formato correto
        output_df['data_iso'] = df["Horário"]

        # Converte a coluna de Velocidade para inteiro, removendo a unidade ' km/h'
        output_df['vel'] = df['Velocidade'].str.replace(' km/h', '').astype(int)

        # Converte latitude e longitude
        output_df['latitude'] = pd.to_numeric(df['Latitude'].str.replace(',', '.', regex=False), errors='coerce').round(5)
        output_df['longitude'] = pd.to_numeric(df['Longitude'].str.replace(',', '.', regex=False), errors='coerce').round(5)

        # Função para extrair informações de endereço do formato Positron
        def extract_positron_address(address):
            """
            Extrai rua, cidade e UF de um endereço Positron.
            """
            try:
                if not address or pd.isna(address):
                    return pd.Series([None, None, None])

                # Divide o endereço por vírgulas
                partes = address.split(',')
                
                # Caso tenha apenas cidade e UF
                if len(partes) == 1 and '-' in partes[0]:
                    cidade_uf = partes[0].split(' - ')
                    cidade = cidade_uf[0].strip()
                    uf = cidade_uf[1].strip() if len(cidade_uf) > 1 else None
                    return pd.Series([None, cidade, uf])
                
                # Caso tenha rua e cidade, mas sem UF
                if len(partes) == 2:
                    rua = partes[0].strip()
                    cidade = partes[1].strip()
                    return pd.Series([rua, cidade, None])
                
                # Caso completo (Rua, Número, Cidade - UF)
                if len(partes) > 2:
                    rua = f"{partes[0].strip()}, {partes[1].strip()}"

                    restante = partes[2].strip()
                    cidade_uf = restante.split(' - ')
                    cidade = cidade_uf[0].strip() if len(cidade_uf) > 0 else None
                    uf = cidade_uf[1].strip() if len(cidade_uf) > 1 else None

                    return pd.Series([rua, cidade, uf])

                # Caso não se encaixe em nenhum padrão
                return pd.Series([address, None, None])

            except Exception as e:
                logger.register_log(f"Erro ao extrair endereço: {e}")
                return pd.Series([None, None, None])

        # Aplica a função de extração de endereço para a coluna 'Endereço'
        output_df[['rua', 'cidade', 'uf']] = df['Endereço'].apply(extract_positron_address)

        # Converte a coluna de ignição para 'Ligada' ou 'Desligada'
        output_df["ignicao"] = df["Ignição"].apply(
            lambda x: "Ligada" if str(x).strip().upper() == "LIGADA" else "Desligada")

    # Caso o sistema de rastreamento seja 'sasgc'
    elif system_type == 'sasgc':
        # Lê o arquivo CSV
        df = pd.read_csv(filepath, encoding='ISO-8859-1', sep=';')

        # Converte a coluna "Data Posicão" para datetime
        temp_data = df["Data Posicão"].str.extract(r'(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2})')[0]
        temp_data = pd.to_datetime(temp_data, dayfirst=True, errors="coerce")

        output_df['data_iso'] = temp_data.dt.strftime('%Y-%m-%d %H:%M:%S')

        # Preenche a coluna de Velocidade
        output_df["vel"] = df["Vel."]

        # Preenche as colunas de Latitude e Longitude
        output_df['latitude'] = df['Latitude']
        output_df['longitude'] = df['Longitude']

        # Preenche as colunas de UF, Cidade e Rua
        output_df['uf'] = df['UF']
        output_df['cidade'] = df['Cidade']
        output_df['rua'] = ''

        # Converte a coluna de Ignicao para 'Ligada' ou 'Desligada'
        output_df["ignicao"] = df["Ign"].apply(lambda x: "Ligada" if str(x).strip().upper() == "SIM" else "Desligada")

    # Caso o sistema de rastreamento seja 'sascar' ou outro
    else:
        # Lê o arquivo Excel
        df = pd.read_excel(filepath)

        # Processa as placas e associa ao id do caminhão (apenas se auto_create_trucks for True)
        plates = df['placa'].str.split('-').str[0]
        unique_plates = plates.unique()
        
        if auto_create_trucks and truck_driver:
            conn = sqlite3.connect(DB_PATH)
            plate_df = pd.read_sql_query("SELECT id, placa FROM trucks", conn)

            df_unique_plates = plate_df['placa'].unique()

            missing_plates = [plate for plate in unique_plates if plate not in df_unique_plates]

            if len(missing_plates) > 0:
                for missing_plate in missing_plates:
                    logger.register_log(f"[INFO] Adicionando placa {missing_plate} ao banco de dados")
                    truck_driver.create_truck(placa=missing_plate)
                plate_df = pd.read_sql_query("SELECT id, placa FROM trucks", conn)

            conn.close()
            plate_to_id_mapping = dict(zip(plate_df['placa'], plate_df['id']))
            output_df['truck_id'] = plates.map(plate_to_id_mapping)
        else:
            # Se não deve criar caminhões automaticamente, deixa truck_id vazio
            output_df['truck_id'] = None

        # Converte a coluna de DataPosicao para datetime
        temp_data = pd.to_datetime(df["dataPosicao"], dayfirst=True, errors='coerce')

        output_df['data_iso'] = temp_data.dt.strftime('%Y-%m-%d %H:%M:%S')

        # Preenche a coluna de Velocidade
        output_df['vel'] = df['velocidade']

        # Converte Latitude e Longitude
        output_df['latitude'] = pd.to_numeric(df['latitude'].str.replace(',', '.', regex=False), errors='coerce')
        output_df['longitude'] = pd.to_numeric(df['longitude'].str.replace(',', '.', regex=False), errors='coerce')

        # Remove linhas com valores inválidos para latitude ou longitude
        output_df = output_df.dropna(subset=['latitude', 'longitude']).reset_index(drop=True)

        # Preenche as colunas de UF, Cidade e Rua
        output_df['uf'] = df['uf']
        output_df['cidade'] = df['cidade']
        output_df['rua'] = df['rua']

        # Converte Ignicao para 'Ligada' ou 'Desligada'
        output_df["ignicao"] = df["ignicao"].apply(lambda x: "Ligada" if str(x).strip() == "1" else "Desligada")

    # Retorna o DataFrame final com os dados extraídos e processados
    return output_df




def fill_excel(name, start, end, tabela, totals, template_path="file_templates/excel_template.xlsx"):
    replace_dict = {
        'motorist_name': name,
        'total_lunch': totals['refeicao'],
        'total_interstice': totals['intersticio'],
        'total_interval': totals['intervalo'],
        'total_load_unload': totals['carga'],
        'total_journey': totals['jornada'],
        'total_driving': totals['direcao'],
        'total_no_stop_driving': totals['sem_pausa'],
        'total_infractions': totals['infracoes'],
        'start': start,
        'end': end
    }

    wb = load_workbook(template_path)
    ws = wb.active

    # Substituir placeholders
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for chave, valor in replace_dict.items():
                    placeholder = f'{{{{{chave}}}}}'
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, str(valor))

    base_line = 11
    current_order = [2, 0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
    line_height_per_wrap = 13
    start_col = 3

    style_template = {}
    for j in range(len(current_order)):
        col = start_col + j
        cell = ws.cell(row=base_line, column=col)
        style_template[col] = {
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "number_format": copy(cell.number_format),
            "protection": copy(cell.protection),
            "alignment": Alignment(wrap_text=True),
        }

    def calculate_required_height(cell_value, column_width, font_size=11):
        """
        Calcula a altura necessária para uma célula baseada no conteúdo e largura da coluna.
        Usa cálculo preciso baseado no tamanho das palavras e largura da fonte.
        
        Args:
            cell_value (str): Conteúdo da célula
            column_width (float): Largura da coluna em caracteres
            font_size (int): Tamanho da fonte
        
        Returns:
            int: Número de linhas necessárias
        """
        if not cell_value or cell_value.strip() == "":
            return 1
        
        # Quebras de linha explícitas
        explicit_lines = cell_value.count('\n') + 1
        
        # Largura da fonte em pixels (aproximação baseada no tamanho da fonte)
        # Fonte 11pt ≈ 8.25 pixels de largura média por caractere
        font_width_px = font_size * 0.75
        
        # Largura da coluna em pixels (aproximação)
        # Excel usa aproximadamente 7 pixels por caractere de largura de coluna
        # Adiciona pequena margem de segurança (5%)
        column_width_px = column_width * 7 * 0.95
        
        def get_char_width(char):
            """Retorna a largura aproximada de um caractere em pixels"""
            if char.isupper():
                return font_width_px * 1.1  # Letras maiúsculas são mais largas
            elif char.isdigit():
                return font_width_px * 0.9  # Números são mais estreitos
            elif char in 'WwMm':
                return font_width_px * 1.3  # Caracteres muito largos
            elif char in 'IiLl':
                return font_width_px * 0.6  # Caracteres muito estreitos
            elif char.isspace():
                return font_width_px * 0.5  # Espaços são mais estreitos
            else:
                return font_width_px  # Largura padrão
        
        # Calcula quantas linhas são necessárias para o texto
        text_lines = 0
        for line in cell_value.split('\n'):
            if line.strip():
                # Calcula a largura total do texto em pixels
                total_text_width = sum(get_char_width(char) for char in line)
                
                # Se o texto cabe em uma linha
                if total_text_width <= column_width_px:
                    text_lines += 1
                else:
                    # Calcula quantas linhas são necessárias
                    # Considera que palavras não são quebradas no meio
                    words = line.split()
                    current_line_width = 0
                    lines_needed = 1
                    
                    for word in words:
                        # Calcula largura da palavra considerando cada caractere
                        word_width = sum(get_char_width(char) for char in word)
                        space_width = get_char_width(' ')  # Espaço entre palavras
                        
                        # Se a palavra cabe na linha atual
                        if current_line_width + word_width <= column_width_px:
                            current_line_width += word_width + space_width
                        else:
                            # Nova linha
                            lines_needed += 1
                            current_line_width = word_width + space_width
                    
                    text_lines += lines_needed
            else:
                text_lines += 1  # Linha vazia conta como uma linha
        
        # Retorna o máximo entre linhas explícitas e calculadas
        return max(explicit_lines, text_lines)

    for i, data_row in enumerate(tabela):
        ordered_row = [data_row[idx] for idx in current_order]
        max_lines = 1
        excel_row = base_line + i

        for col_index in [2, 18]:  # B e R
            base_cell = ws.cell(row=base_line, column=col_index)
            target_cell = ws.cell(row=excel_row, column=col_index)
            target_cell.font = copy(base_cell.font)
            target_cell.fill = copy(base_cell.fill)
            target_cell.border = copy(base_cell.border)
            target_cell.number_format = copy(base_cell.number_format)
            target_cell.protection = copy(base_cell.protection)
            target_cell.alignment = copy(base_cell.alignment)

        for j, value in enumerate(ordered_row):
            col = start_col + j
            cell = ws.cell(row=excel_row, column=col)
            cell_value = str(value) if value is not None else ""
            cell.value = cell_value

            template = style_template[col]
            cell.font = template["font"]
            cell.fill = template["fill"]
            cell.border = template["border"]
            cell.number_format = template["number_format"]
            cell.protection = template["protection"]
            cell.alignment = template["alignment"]

            # Obtém a largura da coluna (mantém a largura padrão do modelo)
            column_letter = cell.column_letter
            column_width = ws.column_dimensions[column_letter].width
            if column_width is None:
                column_width = 10  # Largura padrão se não estiver definida
            
            # Calcula a altura necessária para esta célula
            required_lines = calculate_required_height(cell_value, column_width, 
                                                     font_size=template["font"].size or 11)
            max_lines = max(max_lines, required_lines)

        # Define a altura da linha baseada no conteúdo calculado
        font_size = template["font"].size or 11
        
        # Se precisa de apenas uma linha, usa altura mínima
        if max_lines == 1:
            final_height = max(15, font_size * 1.1)
        else:
            # Para múltiplas linhas, calcula altura baseada no conteúdo
            base_line_height = max(15, font_size * 1.2)
            total_height = max_lines * base_line_height
            # Adiciona pequeno padding apenas para múltiplas linhas
            final_height = total_height * 1.1
        
        # Define limites mínimo e máximo para a altura
        final_height = max(15, min(200, final_height))
        
        ws.row_dimensions[excel_row].height = final_height

    last_row = base_line + len(tabela) - 1
    next_row = last_row + 1
    end_col = start_col + len(current_order) - 1

    left_border_c = ws.cell(row=base_line, column=3).border.left
    left_border_b = ws.cell(row=base_line, column=2).border.left
    bottom_side_c = Side(style=left_border_c.style, color=left_border_c.color)
    bottom_side_b = Side(style=left_border_b.style, color=left_border_b.color)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=last_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=bottom_side_c
        )

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=next_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=bottom_side_b
        )

    cell_b = ws.cell(row=next_row, column=2)
    base_b = ws.cell(row=base_line, column=2).border
    cell_b.border = Border(
        left=Side(style=base_b.left.style, color=base_b.left.color),
        right=cell_b.border.right,
        top=cell_b.border.top,
        bottom=bottom_side_b
    )

    cell_r = ws.cell(row=next_row, column=18)
    base_r = ws.cell(row=base_line, column=18).border
    cell_r.border = Border(
        left=cell_r.border.left,
        right=Side(style=base_r.right.style, color=base_r.right.color),
        top=cell_r.border.top,
        bottom=bottom_side_b
    )

    return wb


def fill_excel_fecham(name, start, end, tabela, totals, template_path="file_templates/excel_template_fecham.xlsx"):
    """
    Versão específica para Fechamento de Ponto.
    - Usa template excel_template_fecham.xlsx
    - Substitui placeholders novos solicitados
    - Remove placeholders antigos não desejados (substitui por string vazia)
    - Mantém a lógica de preenchimento de linhas igual ao fill_excel padrão
    """
    # Placeholders novos
    replace_dict = {
        'motorist_name': name,
        # Novos placeholders pedidos
        'hours_worked': totals.get('h_trab', ''),
        'workload': totals.get('carga_horaria', ''),
        'overtime_50%': totals.get('h_extra50', ''),
        'nighttime_overtime': totals.get('h_e_not', ''),
        'night_additional': totals.get('ad_not', ''),
        'daily_sum': totals.get('diaria', ''),
        'food_help': totals.get('aj_aliment', ''),
        # Período
        'start': start,
        'end': end,
        # Placeholders antigos a excluir -> substituir por vazio
        'total_interstice': '',
        'total_load_unload': '',
        'total_journey': '',
        'total_driving': '',
        'total_no_stop_driving': '',
        'total_infractions': '',
        # Mantidos para compatibilidade se existirem no template, mas não usados aqui
        'total_lunch': totals.get('refeicao', ''),
        'total_interval': totals.get('intervalo', ''),
    }

    wb = load_workbook(template_path)
    ws = wb.active

    # Substituir placeholders nas células
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for chave, valor in replace_dict.items():
                    placeholder = f'{{{{{chave}}}}}'
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, str(valor))

    base_line = 11
    start_col = 3

    # Preencher linhas (ordem idêntica à exibida no modal/relatório PDF)
    for i, data_row in enumerate(tabela):
        # data_row já vem como lista de células na ordem: 
        # [PLACA, DATA, DIA, INÍCIO, ALMOÇO_INICIO+FIM, FIM, INTERVALO, ALMOÇO_DUR, H.TRAB, CARGA, H.EXTRA50%, H.EXTRA100%, H.E.NOT, AD.NOT, DIÁRIA, AJ.ALIMENT., OBS]
        excel_row = base_line + i

        # Bordas externas (colunas B e R) caso existam na linha modelo
        for col_index in [2, 18]:
            base_cell = ws.cell(row=base_line, column=col_index)
            target_cell = ws.cell(row=excel_row, column=col_index)
            target_cell.font = copy(base_cell.font)
            target_cell.fill = copy(base_cell.fill)
            target_cell.border = copy(base_cell.border)
            target_cell.number_format = copy(base_cell.number_format)
            target_cell.protection = copy(base_cell.protection)
            target_cell.alignment = copy(base_cell.alignment)

        # Estilos base por coluna, copiados da linha modelo
        style_template = {}
        for j in range(len(data_row)):
            col = start_col + j
            ref_cell = ws.cell(row=base_line, column=col)
            style_template[col] = {
                "font": copy(ref_cell.font),
                "fill": copy(ref_cell.fill),
                "border": copy(ref_cell.border),
                "number_format": copy(ref_cell.number_format),
                "protection": copy(ref_cell.protection),
                "alignment": Alignment(wrap_text=True),
            }

        # Escrever células
        for j, value in enumerate(data_row):
            col = start_col + j
            cell = ws.cell(row=excel_row, column=col)
            cell_value = "" if value is None else str(value)
            cell.value = cell_value
            tmpl = style_template.get(col)
            if tmpl:
                cell.font = tmpl["font"]
                cell.fill = tmpl["fill"]
                cell.border = tmpl["border"]
                cell.number_format = tmpl["number_format"]
                cell.protection = tmpl["protection"]
                cell.alignment = tmpl["alignment"]

        # Altura mínima da linha
        ws.row_dimensions[excel_row].height = max(ws.row_dimensions[base_line].height or 15, 15)

    # Ajustes de borda inferior da última linha (igual ao fill_excel padrão)
    if tabela:
        last_row = base_line + len(tabela) - 1
        next_row = last_row + 1
        n_cols = max((len(r) for r in tabela), default=0)
        end_col = start_col + max(n_cols - 1, 0)
    else:
        # Sem dados, não há bordas a ajustar
        return wb

    left_border_c = ws.cell(row=base_line, column=3).border.left
    left_border_b = ws.cell(row=base_line, column=2).border.left
    bottom_side_c = Side(style=left_border_c.style, color=left_border_c.color)
    bottom_side_b = Side(style=left_border_b.style, color=left_border_b.color)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=last_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=bottom_side_c
        )

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=next_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=bottom_side_b
        )

    cell_b = ws.cell(row=next_row, column=2)
    base_b = ws.cell(row=base_line, column=2).border
    cell_b.border = Border(
        left=Side(style=base_b.left.style, color=base_b.left.color),
        right=cell_b.border.right,
        top=cell_b.border.top,
        bottom=bottom_side_b
    )

    cell_r = ws.cell(row=next_row, column=18)
    base_r = ws.cell(row=base_line, column=18).border
    cell_r.border = Border(
        left=cell_r.border.left,
        right=Side(style=base_r.right.style, color=base_r.right.color),
        top=cell_r.border.top,
        bottom=bottom_side_b
    )

    return wb

def fill_pdf(name, start, end, tabela, totals):
    """
    Gera um PDF com tabela usando reportlab, com controle preciso sobre o tamanho das linhas.
    
    Args:
        name (str): Nome do motorista
        start (str): Data de início
        end (str): Data de fim
        tabela (list): Lista de dados da tabela
        totals (dict): Dicionário com os totais
    
    Returns:
        bytes: Conteúdo do PDF em bytes
    """
    # Criar buffer para o PDF
    buffer = io.BytesIO()
    
    # Configurar documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Estilo para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=10,
        spaceAfter=5,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    # Estilo para subtítulo
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=10,
        spaceAfter=5,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    # Estilo para parágrafos
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=0,
        alignment=TA_LEFT
    )
    
    # Estilo para célula de infrações
    infraction_style = ParagraphStyle(
        'InfractionCell',
        parent=styles['Normal'],
        fontSize=6,  # Diminuída de 7 para 6
        alignment=TA_LEFT,
        wordWrap='CJK',
        leading=8,  # Diminuída de 9 para 8
        spaceAfter=0,
        spaceBefore=0,
    )
    
    # Elementos do PDF
    elements = []
    
    # Título
    title = Paragraph(f"Relatório de Jornada - {name}", title_style)
    elements.append(title)
    
    # Totais e período em uma linha só
    totals_data = [
        ["Período", "Tempo de Refeição", "Interstício", "Tempo de Intervalo", "Tempo de C/D", "Jornada Total", "Tempo de Direção", "Direção sem Pausa", "Infrações"],
        [f"{start} a {end}", totals['refeicao'], totals['intersticio'], totals['intervalo'], totals['carga'], totals['jornada'], totals['direcao'], totals['sem_pausa'], totals['infracoes']]
    ]
    
    # Criar tabela de totais compacta
    totals_table = Table(totals_data, colWidths=[120, 80, 60, 80, 80, 80, 80, 80, 60], spaceAfter=5)
    
    # Estilo da tabela de totais compacta
    totals_style = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 2),  # Diminuída de 4 para 3
        ('TOPPADDING', (0, 0), (-1, 0), 2),  # Diminuída de 4 para 3
        
        # Valores
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, 1), 8),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
        
        # Padding mínimo
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        
        # Bordas
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ])
    
    totals_table.setStyle(totals_style)
    elements.append(totals_table)
    
    #elements.append(Spacer(1, 5))  # Reduzido de 20 para 15

    # Preparar dados da tabela
    if tabela:
        # Cabeçalhos com quebras de linha (removida coluna observação)
        headers = [
            "Data", "Dia", "Placa", "Início\nJornada", "In.\nRefeição", "Fim\nRefeição",
            "Fim\nJornada", "Tempo\nRefeição", "Interstício", "Tempo\nIntervalo",
            "Tempo\nC/D", "Jornada\nTotal", "Tempo\nDireção", "Direção\ns/ Pausa", "Infrações"
        ]
        
        # Dados da tabela
        table_data = [headers]
        
        # Ordem das colunas (ordem correta da página, sem observação)
        current_order = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
        
        for row in tabela:
            # Garante que todos os campos são string
            row = [str(col) if col is not None else "" for col in row]
            # Reorganizar dados na ordem correta
            ordered_row = [row[idx] for idx in current_order]
            # Processar texto das infrações para quebrar linha
            infractions_text = str(row[14]).replace('|', '<br/>')
            infractions_text = infractions_text.replace(';', '; ')
            infractions_text = re.sub(r'\([^)]*\)', '', infractions_text)
            infractions_text = re.sub(r'\s+', ' ', infractions_text).strip()
            ordered_row.pop(14)
            ordered_row.append(Paragraph(infractions_text, infraction_style))
            table_data.append(ordered_row)
        
        # Calcular larguras das colunas (em pontos)
        # A4 landscape: 842 pontos de largura, menos margens = ~780 pontos
        available_width = 780
        num_cols = len(headers)
        
        # Larguras específicas para cada coluna (em pontos) - removida observação, aumentada infrações
        column_widths = [
            50,   # Data
            50,   # Dia
            60,   # Placa
            40,   # Início Jornada (só horário HH:mm)
            40,   # In. Refeição (só horário HH:mm)
            40,   # Fim Refeição (só horário HH:mm)
            40,   # Fim Jornada (só horário HH:mm)
            40,   # Tempo Refeição (só horário HH:mm)
            50,   # Interstício
            40,   # Tempo Intervalo (só horário HH:mm)
            40,   # Tempo C/D (só horário HH:mm)
            40,   # Jornada Total (só horário HH:mm)
            40,   # Tempo Direção (só horário HH:mm)
            50,   # Direção s/ Pausa
            370   # Infrações 
        ]
        
        # Ajustar larguras para caber na página
        total_width = sum(column_widths)
        if total_width > available_width:
            scale_factor = available_width / total_width
            column_widths = [w * scale_factor for w in column_widths]
        
        # Criar tabela
        table = Table(table_data, colWidths=column_widths)
        
        # Estilo da tabela
        table_style = TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),  # Diminuída de 4 para 3
            ('TOPPADDING', (0, 0), (-1, 0), 2),  # Diminuída de 4 para 3
            ('LEFTPADDING', (0, 0), (-1, 0), 4),
            ('RIGHTPADDING', (0, 0), (-1, 0), 4),
            
            # Dados - todas as colunas centralizadas exceto infrações
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('ALIGN', (0, 1), (-2, -1), 'CENTER'),  # Todas exceto última coluna
            ('ALIGN', (-1, 1), (-1, -1), 'LEFT'),   # Última coluna (infrações) alinhada à esquerda
            ('BOTTOMPADDING', (0, 1), (-1, -1), 0),  # Diminuída de 2 para 1
            ('TOPPADDING', (0, 1), (-1, -1), 0),  # Diminuída de 2 para 1
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            
            # Bordas
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            
            # Quebra de linha para cabeçalhos e dados
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
    
    # Gerar PDF
    doc.build(elements)
    
    # Retornar bytes
    buffer.seek(0)
    return buffer.getvalue()


def calculate_pdf_row_height(cell_value, column_width, font_size=6):
    """
    Calcula a altura necessária para uma linha na tabela PDF.
    
    Args:
        cell_value (str): Conteúdo da célula
        column_width (float): Largura da coluna em pontos
        font_size (int): Tamanho da fonte
    
    Returns:
        float: Altura necessária em pontos
    """
    if not cell_value or str(cell_value).strip() == "":
        return font_size * 1.0  # Altura mínima reduzida de 1.2 para 1.0
    
    # Largura aproximada de um caractere em pontos
    char_width = font_size * 0.6
    
    # Número de caracteres que cabem na coluna
    chars_per_line = max(1, int(column_width / char_width))
    
    # Quebras de linha explícitas
    explicit_lines = str(cell_value).count('\n') + 1
    
    # Calcular linhas necessárias para o texto
    text_lines = 0
    for line in str(cell_value).split('\n'):
        if line.strip():
            # Se o texto cabe em uma linha
            if len(line) <= chars_per_line:
                text_lines += 1
            else:
                # Calcular quantas linhas são necessárias
                lines_needed = math.ceil(len(line) / chars_per_line)
                text_lines += lines_needed
        else:
            text_lines += 1
    
    # Altura total necessária
    total_lines = max(explicit_lines, text_lines)
    row_height = total_lines * font_size * 1.0  # Diminuída de 1.1 para 1.0
    
    # Limites mínimo e máximo
    min_height = font_size * 1.0  # Diminuída de 1.2 para 1.0
    max_height = font_size * 6  # Diminuída de 8 para 6
    
    return max(min_height, min(max_height, row_height))
